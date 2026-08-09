import io
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime

import pytest
import requests
from openpyxl import load_workbook
from conftest import load_service

pytestmark = pytest.mark.integration
doc_module = load_service("document_service_integration", "services/document_service/app.py")


def make_payload(base, doc_type, suffix=""):
    payload = dict(base)
    payload["asunto"] += suffix
    if doc_type == "informes": payload["tipo_informe"] = "Informe técnico"
    if doc_type == "reportes": payload["tipo_reporte"] = "Reporte de seguimiento"
    return payload


def test_create_read_update_list_and_delete_all_document_types(urls, admin_headers, base_document_payload):
    created = {}
    for doc_type in ("actas","informes","reportes","comisiones"):
        response = requests.post(f"{urls['documents']}/api/documents/{doc_type}", headers=admin_headers, json=make_payload(base_document_payload, doc_type, doc_type), timeout=20)
        assert response.status_code == 201, f"{doc_type}: {response.text}"
        item = response.json(); created[doc_type] = item
        expected_prefix = {"actas":"ACTAS.DTCD","informes":"INF.DTCD","reportes":"REP.DTCD","comisiones":"CMS.DTCD"}[doc_type]
        assert item["code"].startswith(expected_prefix + ".") and item["code"].endswith(f".{datetime.now().year}")
        fetched = requests.get(f"{urls['documents']}/api/documents/{doc_type}/{item['id']}", headers=admin_headers, timeout=15)
        assert fetched.status_code == 200 and fetched.json()["subject"] == item["subject"]
        updated_payload = make_payload(base_document_payload, doc_type, " actualizado")
        updated_payload["observaciones"] = "Observación actualizada"
        updated = requests.put(f"{urls['documents']}/api/documents/{doc_type}/{item['id']}", headers=admin_headers, json=updated_payload, timeout=20)
        assert updated.status_code == 200 and updated.json()["observations"] == "Observación actualizada"
    listing = requests.get(f"{urls['documents']}/api/documents/actas?per_page=all", headers=admin_headers, timeout=15)
    assert listing.status_code == 200 and any(x["id"] == created["actas"]["id"] for x in listing.json()["items"])
    deleted = requests.delete(f"{urls['documents']}/api/documents/comisiones/{created['comisiones']['id']}", headers=admin_headers, timeout=15)
    assert deleted.status_code == 200


def test_all_users_can_view_all_documents_but_only_owner_can_edit(urls, admin_headers, normal_headers, normal_user, create_user, base_document_payload):
    second = create_user("segundo")
    second_headers = {"X-Internal-Key":admin_headers["X-Internal-Key"],"X-User-ID":str(second["id"]),"X-User-Role":"usuario"}
    own = requests.post(f"{urls['documents']}/api/documents/actas", headers=normal_headers, json={**base_document_payload, "asunto": "Documento visible del primer usuario"}, timeout=20)
    other = requests.post(f"{urls['documents']}/api/documents/actas", headers=second_headers, json={**base_document_payload, "asunto": "Documento visible del segundo usuario"}, timeout=20)
    assert own.status_code == 201 and other.status_code == 201
    own_id = own.json()["id"]
    other_id = other.json()["id"]
    assert requests.get(f"{urls['documents']}/api/documents/actas/{own_id}", headers=second_headers, timeout=15).status_code == 200
    assert requests.get(f"{urls['documents']}/api/documents/actas/{other_id}", headers=normal_headers, timeout=15).status_code == 200
    listing = requests.get(f"{urls['documents']}/api/documents/actas?per_page=all", headers=normal_headers, timeout=15)
    listed_ids = {item["id"] for item in listing.json()["items"]}
    assert listing.status_code == 200 and {own_id, other_id}.issubset(listed_ids)
    assert requests.put(f"{urls['documents']}/api/documents/actas/{own_id}", headers=second_headers, json=base_document_payload, timeout=15).status_code == 403
    assert requests.delete(f"{urls['documents']}/api/documents/actas/{own_id}", headers=normal_headers, timeout=15).status_code == 403


def test_dynamic_jsonb_answers_and_historical_values_survive_edit(urls, admin_headers, base_document_payload):
    payload = dict(base_document_payload)
    payload["custom_fields"] = {"numero_contrato":"CTR-001","estado":"Otros","estado__other":"En revisión","campo_historico":"No borrar"}
    payload["form_definition_snapshot"] = [
        {"field_key":"numero_contrato","label":"Número de contrato"},
        {"field_key":"estado","label":"Estado"},
        {"field_key":"campo_historico","label":"Campo histórico"},
    ]
    created = requests.post(f"{urls['documents']}/api/documents/actas", headers=admin_headers, json=payload, timeout=20).json()
    edited = dict(base_document_payload)
    edited["custom_fields"] = {"numero_contrato":"CTR-002"}
    edited["form_definition_snapshot"] = [{"field_key":"numero_contrato","label":"Código de contrato"}]
    response = requests.put(f"{urls['documents']}/api/documents/actas/{created['id']}", headers=admin_headers, json=edited, timeout=20)
    assert response.status_code == 200, response.text
    extra = response.json()["extra_data"]
    assert extra["numero_contrato"] == "CTR-002"
    assert extra["campo_historico"] == "No borrar"


def test_annual_counter_is_transactional_and_resets_per_year(document_db):
    with document_db.cursor() as cur:
        first = doc_module.allocate_number(cur, "reportes", 2098)
        second = doc_module.allocate_number(cur, "reportes", 2098)
        new_year = doc_module.allocate_number(cur, "reportes", 2099)
        other_type = doc_module.allocate_number(cur, "actas", 2098)
        assert (first, second, new_year, other_type) == (1, 2, 1, 1)
    document_db.rollback()


def test_concurrent_creation_never_duplicates_codes(urls, admin_headers, base_document_payload):
    def create_one(index):
        payload = make_payload(base_document_payload, "reportes", f" concurrente {index}")
        response = requests.post(f"{urls['documents']}/api/documents/reportes", headers=admin_headers, json=payload, timeout=30)
        return response.status_code, response.text, response.json() if response.status_code == 201 else None
    with ThreadPoolExecutor(max_workers=8) as executor:
        results = list(executor.map(create_one, range(12)))
    assert all(status == 201 for status, _, _ in results), results
    codes = [item["code"] for _, _, item in results]
    assert len(codes) == len(set(codes))


def test_excel_export_contains_dynamic_columns_and_neutralizes_formulas(urls, admin_headers, base_document_payload):
    payload = dict(base_document_payload)
    payload["asunto"] = "=HYPERLINK(\"https://malicioso\")"
    payload["custom_fields"] = {"campo_excel":"=SUM(1,1)"}
    payload["form_definition_snapshot"] = [{"field_key":"campo_excel","label":"Campo Excel"}]
    created = requests.post(f"{urls['documents']}/api/documents/actas", headers=admin_headers, json=payload, timeout=20)
    assert created.status_code == 201
    exported = requests.get(f"{urls['documents']}/api/documents/actas/export.xlsx", headers=admin_headers, timeout=30)
    assert exported.status_code == 200
    workbook = load_workbook(io.BytesIO(exported.content), data_only=False)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    assert "Campo adicional: Campo Excel" in headers
    subject_col = headers.index("Asunto") + 1
    extra_col = headers.index("Campo adicional: Campo Excel") + 1
    rows = list(sheet.iter_rows(min_row=2))
    target = next(row for row in rows if row[subject_col-1].value and "HYPERLINK" in str(row[subject_col-1].value))
    assert str(target[subject_col-1].value).startswith("'=")
    assert str(target[extra_col-1].value).startswith("'=")


def test_document_audit_log_is_written(document_db):
    with document_db.cursor() as cur:
        cur.execute("SELECT COUNT(*) FROM audit_log WHERE action IN ('DOCUMENT_CREATE','DOCUMENT_UPDATE','DOCUMENT_DELETE')")
        assert cur.fetchone()[0] > 0


def test_manual_number_is_admin_only_and_duplicates_are_rejected(urls, admin_headers, normal_headers, base_document_payload):
    manual = 8000
    first = requests.post(f"{urls['documents']}/api/documents/comisiones", headers=admin_headers,
        json={**base_document_payload,"numero_manual":manual}, timeout=20)
    assert first.status_code == 201, first.text
    assert first.json()["number"] == manual
    duplicate = requests.post(f"{urls['documents']}/api/documents/comisiones", headers=admin_headers,
        json={**base_document_payload,"numero_manual":manual}, timeout=20)
    assert duplicate.status_code in {400,409}
    forbidden = requests.post(f"{urls['documents']}/api/documents/comisiones", headers=normal_headers,
        json={**base_document_payload,"numero_manual":manual+1}, timeout=20)
    assert forbidden.status_code == 403


def test_normal_user_excel_export_contains_records_from_all_users(urls, normal_headers, normal_user, create_user, base_document_payload):
    other = create_user("excelotro")
    other_headers = {"X-Internal-Key":normal_headers["X-Internal-Key"],"X-User-ID":str(other["id"]),"X-User-Role":"usuario"}
    own_subject = "Exportación propia única"
    other_subject = "Exportación ajena única"
    requests.post(f"{urls['documents']}/api/documents/actas", headers=normal_headers, json={**base_document_payload,"asunto":own_subject}, timeout=20)
    requests.post(f"{urls['documents']}/api/documents/actas", headers=other_headers, json={**base_document_payload,"asunto":other_subject}, timeout=20)
    exported = requests.get(f"{urls['documents']}/api/documents/actas/export.xlsx", headers=normal_headers, timeout=30)
    assert exported.status_code == 200
    workbook = load_workbook(io.BytesIO(exported.content), data_only=False)
    sheet = workbook.active
    values = [[cell.value for cell in row] for row in sheet.iter_rows()]
    flattened = " ".join(str(value) for row in values for value in row if value is not None)
    assert own_subject in flattened
    assert other_subject in flattened
