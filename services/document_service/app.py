import io
import math
import os
import re
import secrets
from datetime import date, datetime, time
from zoneinfo import ZoneInfo

import psycopg2
import requests
from flask import Flask, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from psycopg2.extras import Json, RealDictCursor

app = Flask(__name__)
DATABASE_URL = os.environ["DATABASE_URL"]
INTERNAL_API_KEY = os.environ["INTERNAL_API_KEY"]
if len(INTERNAL_API_KEY) < 32:
    raise RuntimeError("INTERNAL_API_KEY debe tener al menos 32 caracteres")
APP_TZ = os.getenv("TZ", "America/Guayaquil")
AUTH_SERVICE_URL = os.getenv("AUTH_SERVICE_URL", "http://auth-service:8000")
NOTIFICATION_SERVICE_URL = os.getenv("NOTIFICATION_SERVICE_URL", "http://notification-service:8000")
CATALOG_SERVICE_URL = os.getenv("CATALOG_SERVICE_URL", "http://catalog-service:8000")
ALLOWED_TYPES = {"actas", "informes", "reportes", "comisiones"}
PREFIXES = {"actas": "ACTAS.DTCD", "informes": "INF.DTCD", "reportes": "REP.DTCD", "comisiones": "CMS.DTCD"}


def json_ready(value):
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, dict):
        return {key: json_ready(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [json_ready(item) for item in value]
    return value


def conn():
    return psycopg2.connect(DATABASE_URL)


def require_internal(admin=False):
    if not secrets.compare_digest(request.headers.get("X-Internal-Key", ""), INTERNAL_API_KEY):
        return jsonify(error="Acceso interno no autorizado"), 401
    if admin and request.headers.get("X-User-Role") != "admin":
        return jsonify(error="Se requiere rol administrador"), 403
    return None


def current_user_id():
    try:
        return int(request.headers.get("X-User-ID", "0"))
    except ValueError:
        return 0


def clean_text(value, field, max_len, required=False):
    text = str(value or "").strip()
    text = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", text)
    if required and not text:
        raise ValueError(f"{field} es obligatorio")
    if len(text) > max_len:
        raise ValueError(f"{field} supera el máximo de {max_len} caracteres")
    return text or None


def code_for(doc_type, year, number):
    return f"{PREFIXES[doc_type]}.{int(number):03d}.{int(year)}"


def internal_headers(role="admin", user_id=None):
    headers = {"X-Internal-Key": INTERNAL_API_KEY, "X-User-Role": role}
    if user_id:
        headers["X-User-ID"] = str(user_id)
    return headers


def get_user(user_id):
    try:
        response = requests.get(
            f"{AUTH_SERVICE_URL}/api/users/{int(user_id)}",
            headers=internal_headers(),
            timeout=(3, 10),
        )
        return response.json() if response.status_code == 200 else None
    except (requests.RequestException, ValueError):
        app.logger.exception("No se pudo consultar el usuario %s", user_id)
        return None


def get_custom_labels(doc_type):
    try:
        response = requests.get(
            f"{CATALOG_SERVICE_URL}/api/form-fields/{doc_type}?include_inactive=1",
            headers=internal_headers(),
            timeout=(3, 10),
        )
        if response.status_code == 200:
            return {item["field_key"]: item["label"] for item in response.json().get("items", [])}
    except requests.RequestException:
        app.logger.exception("No se pudieron consultar los nombres de campos")
    return {}


def notify_document(event_type, item):
    user = get_user(item.get("user_id"))
    if not user or not user.get("email"):
        return
    context = dict(json_ready(item))
    context["owner_name"] = user.get("nombre")
    context["document_label"] = {
        "actas": "Acta", "informes": "Informe", "reportes": "Reporte", "comisiones": "Comisión"
    }.get(item.get("document_type"), item.get("document_type"))
    context["custom_labels"] = get_custom_labels(item.get("document_type"))
    try:
        response = requests.post(
            f"{NOTIFICATION_SERVICE_URL}/api/notifications",
            headers=internal_headers(),
            json={"event_type": event_type, "recipient": user["email"], "context": context},
            timeout=(4, 25),
        )
        if response.status_code >= 400:
            app.logger.warning("Notificación rechazada: %s", response.text[:500])
    except requests.RequestException:
        app.logger.exception("No se pudo registrar la notificación documental")

def allocate_number(cur, doc_type, year, manual=None, exclude_id=None):
    if manual not in (None, ""):
        number = int(manual)
        if number <= 0:
            raise ValueError("El número correlativo debe ser mayor que cero")
        if exclude_id is None:
            cur.execute(
                "SELECT 1 FROM documents WHERE document_type=%s AND year=%s AND number=%s",
                (doc_type, year, number),
            )
        else:
            cur.execute(
                "SELECT 1 FROM documents WHERE document_type=%s AND year=%s AND number=%s AND id<>%s",
                (doc_type, year, number, exclude_id),
            )
        if cur.fetchone():
            raise ValueError("Ese número correlativo ya existe para el año indicado")
        cur.execute(
            """
            INSERT INTO document_counters(document_type,year,next_number)
            VALUES(%s,%s,%s)
            ON CONFLICT(document_type,year)
            DO UPDATE SET next_number=GREATEST(document_counters.next_number,EXCLUDED.next_number)
            """,
            (doc_type, year, number + 1),
        )
        return number
    cur.execute(
        """
        INSERT INTO document_counters(document_type,year,next_number)
        VALUES(%s,%s,2)
        ON CONFLICT(document_type,year)
        DO UPDATE SET next_number=document_counters.next_number+1
        RETURNING next_number-1 AS allocated_number
        """,
        (doc_type, year),
    )
    row = cur.fetchone()
    # Funciona tanto con cursor normal (tupla) como con RealDictCursor.
    return row["allocated_number"] if isinstance(row, dict) else row[0]


def validate_payload(doc_type, data):
    if doc_type not in ALLOWED_TYPES:
        raise ValueError("Tipo de documento inválido")
    payload = {
        "company": clean_text(data.get("empresa"), "La empresa", 200, True),
        "management": clean_text(data.get("gestiones"), "La gestión", 500, True),
        "associated_products": clean_text(data.get("productos_asociados"), "El producto asociado", 1000, True),
        "subject": clean_text(data.get("asunto"), "El asunto", 255, True),
        "observations": clean_text(data.get("observaciones"), "Las observaciones", 5000),
        "subtype": None,
        "case_type": None,
        "feeder_name": None,
        "substation": None,
        "subtransmission_line": None,
        "interruption_date": None,
    }
    if doc_type == "informes":
        payload["subtype"] = clean_text(data.get("tipo_informe"), "El tipo de informe", 300, True)
        if "caso fortuito" in payload["subtype"].lower():
            # Los requisitos del caso fortuito se administran desde el constructor visual.
            # Estas columnas se mantienen para compatibilidad con documentos anteriores.
            payload["case_type"] = clean_text(data.get("caso_tipo"), "El elemento afectado", 60)
            if payload["case_type"] and payload["case_type"] not in {"ALIMENTADOR", "LINEAS DE SUBTRANSMISION"}:
                raise ValueError("Elemento afectado inválido")
            interruption = clean_text(data.get("fecha_interrupcion"), "La fecha de interrupción", 10)
            if interruption:
                try:
                    datetime.strptime(interruption, "%Y-%m-%d")
                except ValueError as exc:
                    raise ValueError("La fecha de interrupción no es válida") from exc
                payload["interruption_date"] = interruption
            if payload["case_type"] == "ALIMENTADOR":
                payload["feeder_name"] = clean_text(data.get("nombre_alimentador"), "El alimentador", 200)
                payload["substation"] = clean_text(data.get("alimentador_subestacion"), "La subestación", 200)
            elif payload["case_type"] == "LINEAS DE SUBTRANSMISION":
                payload["subtransmission_line"] = clean_text(data.get("linea_subtransmision_nombre"), "La línea de subtransmisión", 200)
    elif doc_type == "reportes":
        payload["subtype"] = clean_text(data.get("tipo_reporte"), "El tipo de reporte", 300, True)
    return payload



def validate_custom_fields(raw):
    if raw in (None, ""):
        return {}
    if not isinstance(raw, dict):
        raise ValueError("Los campos adicionales deben enviarse como un objeto")
    if len(raw) > 100:
        raise ValueError("No se permiten más de 100 campos adicionales")
    cleaned = {}
    for key, value in raw.items():
        key = str(key).strip().lower()
        if not re.fullmatch(r"[a-z][a-z0-9_]{0,79}", key):
            raise ValueError("Se recibió una clave de campo adicional inválida")
        if isinstance(value, bool):
            cleaned[key] = value
        elif value is None:
            cleaned[key] = ""
        elif isinstance(value, (int, float)):
            cleaned[key] = value
        elif isinstance(value, list):
            if len(value) > 100:
                raise ValueError(f"El campo {key} contiene demasiadas opciones")
            cleaned[key] = [clean_text(item, f"El campo {key}", 500, True) for item in value]
        else:
            text = clean_text(value, f"El campo {key}", 5000)
            cleaned[key] = text or ""
    return cleaned


def init_db():
    schema = """
    CREATE TABLE IF NOT EXISTS document_counters (
      document_type VARCHAR(20) NOT NULL CHECK(document_type IN ('actas','informes','reportes','comisiones')),
      year INTEGER NOT NULL CHECK(year >= 2020),
      next_number INTEGER NOT NULL CHECK(next_number > 0),
      PRIMARY KEY(document_type,year)
    );
    CREATE TABLE IF NOT EXISTS documents (
      id BIGSERIAL PRIMARY KEY,
      document_type VARCHAR(20) NOT NULL CHECK(document_type IN ('actas','informes','reportes','comisiones')),
      number INTEGER NOT NULL CHECK(number > 0),
      year INTEGER NOT NULL CHECK(year >= 2020),
      code VARCHAR(50) NOT NULL UNIQUE,
      user_id INTEGER NOT NULL,
      company VARCHAR(200) NOT NULL,
      management VARCHAR(500) NOT NULL,
      associated_products VARCHAR(1000) NOT NULL,
      subtype VARCHAR(300),
      subject VARCHAR(255) NOT NULL,
      observations TEXT,
      case_type VARCHAR(60),
      feeder_name VARCHAR(200),
      substation VARCHAR(200),
      subtransmission_line VARCHAR(200),
      interruption_date DATE,
      document_date DATE NOT NULL,
      document_time TIME NOT NULL,
      created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
      updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
      legacy_table VARCHAR(20),
      legacy_id BIGINT,
      extra_data JSONB NOT NULL DEFAULT '{}'::jsonb,
      form_definition_snapshot JSONB NOT NULL DEFAULT '[]'::jsonb,
      UNIQUE(document_type,year,number),
      UNIQUE(legacy_table,legacy_id)
    );
    ALTER TABLE documents ADD COLUMN IF NOT EXISTS legacy_table VARCHAR(20);
    ALTER TABLE documents ADD COLUMN IF NOT EXISTS legacy_id BIGINT;
    ALTER TABLE documents ADD COLUMN IF NOT EXISTS extra_data JSONB NOT NULL DEFAULT '{}'::jsonb;
    ALTER TABLE documents ADD COLUMN IF NOT EXISTS form_definition_snapshot JSONB NOT NULL DEFAULT '[]'::jsonb;
    CREATE UNIQUE INDEX IF NOT EXISTS uq_documents_legacy ON documents(legacy_table,legacy_id) WHERE legacy_table IS NOT NULL;
    CREATE INDEX IF NOT EXISTS idx_documents_type_date ON documents(document_type, document_date DESC);
    CREATE INDEX IF NOT EXISTS idx_documents_user ON documents(user_id);
    CREATE INDEX IF NOT EXISTS idx_documents_company ON documents(company);
    CREATE INDEX IF NOT EXISTS idx_documents_year_number ON documents(year,number);
    CREATE TABLE IF NOT EXISTS audit_log (
      id BIGSERIAL PRIMARY KEY,
      actor_user_id INTEGER,
      action VARCHAR(80) NOT NULL,
      document_id BIGINT,
      detail JSONB,
      created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
    );
    """
    with conn() as db, db.cursor() as cur:
        cur.execute(schema)
        cur.execute(
            """
            UPDATE documents
            SET code = CASE document_type
                WHEN 'actas' THEN 'ACTAS.DTCD'
                WHEN 'informes' THEN 'INF.DTCD'
                WHEN 'reportes' THEN 'REP.DTCD'
                WHEN 'comisiones' THEN 'CMS.DTCD'
              END || '.' || LPAD(number::text,3,'0') || '.' || year::text
            WHERE code <> CASE document_type
                WHEN 'actas' THEN 'ACTAS.DTCD'
                WHEN 'informes' THEN 'INF.DTCD'
                WHEN 'reportes' THEN 'REP.DTCD'
                WHEN 'comisiones' THEN 'CMS.DTCD'
              END || '.' || LPAD(number::text,3,'0') || '.' || year::text
            """
        )
        db.commit()


@app.get("/health")
def health():
    try:
        with conn() as db, db.cursor() as cur:
            cur.execute("SELECT 1")
            cur.fetchone()
        return jsonify(status="ok")
    except Exception as exc:
        return jsonify(status="error", detail=str(exc)), 503


@app.get("/api/documents/companies")
def companies():
    denied = require_internal()
    if denied:
        return denied
    with conn() as db, db.cursor() as cur:
        cur.execute("SELECT DISTINCT company FROM documents WHERE company<>'' ORDER BY company")
        values = [row[0] for row in cur.fetchall()]
    return jsonify(items=values)


@app.get("/api/documents/<doc_type>")
def list_documents(doc_type):
    denied = require_internal()
    if denied:
        return denied
    if doc_type not in ALLOWED_TYPES:
        return jsonify(error="Tipo de documento inválido"), 400
    company = request.args.get("company", "").strip()
    user_id = request.args.get("user_id", type=int)
    if request.headers.get("X-User-Role") != "admin" and current_user_id() <= 0:
        return jsonify(error="Usuario inválido"), 400
    page = max(request.args.get("page", 1, type=int), 1)
    per_page_raw = request.args.get("per_page", "10")
    try:
        per_page = 100000 if per_page_raw == "all" else min(max(int(per_page_raw or 10), 1), 200)
    except ValueError:
        return jsonify(error="Cantidad por página inválida"), 400
    conditions = ["document_type=%s"]
    params = [doc_type]
    if company:
        conditions.append("company=%s")
        params.append(company)
    if user_id:
        conditions.append("user_id=%s")
        params.append(user_id)
    where = " AND ".join(conditions)
    with conn() as db, db.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(f"SELECT COUNT(*) AS total FROM documents WHERE {where}", params)
        total = cur.fetchone()["total"]
        cur.execute(
            f"SELECT * FROM documents WHERE {where} ORDER BY year DESC,number DESC LIMIT %s OFFSET %s",
            params + [per_page, (page - 1) * per_page],
        )
        items = cur.fetchall()
    return jsonify(items=json_ready(items), total=total, pages=max(math.ceil(total / per_page), 1), page=page, per_page=per_page_raw)


@app.get("/api/documents/<doc_type>/<int:doc_id>")
def get_document(doc_type, doc_id):
    denied = require_internal()
    if denied:
        return denied
    if doc_type not in ALLOWED_TYPES:
        return jsonify(error="Tipo de documento inválido"), 400
    with conn() as db, db.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute("SELECT * FROM documents WHERE id=%s AND document_type=%s", (doc_id, doc_type))
        item = cur.fetchone()
    if not item:
        return jsonify(error="Documento no encontrado"), 404
    if request.headers.get("X-User-Role") != "admin" and current_user_id() <= 0:
        return jsonify(error="Usuario inválido"), 400
    return jsonify(json_ready(item))


@app.post("/api/documents/<doc_type>")
def create_document(doc_type):
    denied = require_internal()
    if denied:
        return denied
    try:
        data = request.get_json(silent=True) or {}
        payload = validate_payload(doc_type, data)
        custom_fields = validate_custom_fields(data.get("custom_fields"))
        form_snapshot = data.get("form_definition_snapshot") or []
        if not isinstance(form_snapshot, list):
            raise ValueError("La definición del formulario no es válida")
        user_id = current_user_id()
        if user_id <= 0:
            raise ValueError("Usuario inválido")
        now = datetime.now(ZoneInfo(APP_TZ))
        year = now.year
        manual_number = data.get("numero_manual")
        if manual_number not in (None, "") and request.headers.get("X-User-Role") != "admin":
            return jsonify(error="Solo un administrador puede asignar el correlativo manualmente"), 403
        with conn() as db, db.cursor(cursor_factory=RealDictCursor) as cur:
            number = allocate_number(cur, doc_type, year, manual_number)
            code = code_for(doc_type, year, number)
            cur.execute(
                """
                INSERT INTO documents(document_type,number,year,code,user_id,company,management,associated_products,subtype,subject,observations,case_type,feeder_name,substation,subtransmission_line,interruption_date,document_date,document_time,extra_data,form_definition_snapshot)
                VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                RETURNING *
                """,
                (doc_type, number, year, code, user_id, payload["company"], payload["management"], payload["associated_products"], payload["subtype"], payload["subject"], payload["observations"], payload["case_type"], payload["feeder_name"], payload["substation"], payload["subtransmission_line"], payload["interruption_date"], now.date(), now.time().replace(tzinfo=None), Json(custom_fields), Json(form_snapshot)),
            )
            item = cur.fetchone()
            cur.execute("INSERT INTO audit_log(actor_user_id,action,document_id,detail) VALUES(%s,'DOCUMENT_CREATE',%s,jsonb_build_object('type',%s))", (user_id, item["id"], doc_type))
            db.commit()
        notify_document("document_created", item)
        return jsonify(json_ready(item)), 201
    except (ValueError, TypeError) as exc:
        return jsonify(error=str(exc)), 400
    except psycopg2.errors.UniqueViolation:
        return jsonify(error="El correlativo o código ya existe"), 409
    except psycopg2.Error:
        app.logger.exception("Error de base de datos al crear el documento")
        return jsonify(error="No se pudo guardar el documento"), 500
    except Exception:
        app.logger.exception("Error inesperado al crear el documento")
        return jsonify(error="No se pudo guardar el documento"), 500


@app.put("/api/documents/<doc_type>/<int:doc_id>")
def update_document(doc_type, doc_id):
    denied = require_internal()
    if denied:
        return denied
    try:
        data = request.get_json(silent=True) or {}
        payload = validate_payload(doc_type, data)
        custom_fields = validate_custom_fields(data.get("custom_fields")) if "custom_fields" in data else None
        form_snapshot = data.get("form_definition_snapshot") or []
        if not isinstance(form_snapshot, list):
            raise ValueError("La definición del formulario no es válida")
        actor = current_user_id()
        role = request.headers.get("X-User-Role", "")
        with conn() as db, db.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT * FROM documents WHERE id=%s AND document_type=%s FOR UPDATE", (doc_id, doc_type))
            current = cur.fetchone()
            if not current:
                return jsonify(error="Documento no encontrado"), 404
            if role != "admin" and current["user_id"] != actor:
                return jsonify(error="No tiene permiso para editar este documento"), 403
            year = current["year"]
            number = current["number"]
            if data.get("numero_manual") not in (None, "", str(number), number):
                if role != "admin":
                    return jsonify(error="Solo un administrador puede cambiar el correlativo"), 403
                number = allocate_number(cur, doc_type, year, data.get("numero_manual"), exclude_id=doc_id)
            code = code_for(doc_type, year, number)
            merged_extra = dict(current.get("extra_data") or {})
            if custom_fields is not None:
                active_keys = set()
                for definition in form_snapshot:
                    key = str((definition or {}).get("field_key", "")).strip().lower()
                    if re.fullmatch(r"[a-z][a-z0-9_]{0,79}", key):
                        active_keys.add(key)
                        active_keys.add(f"{key}__other")
                for key in active_keys:
                    merged_extra.pop(key, None)
                merged_extra.update(custom_fields)
            cur.execute(
                """
                UPDATE documents SET number=%s,code=%s,company=%s,management=%s,associated_products=%s,subtype=%s,subject=%s,observations=%s,case_type=%s,feeder_name=%s,substation=%s,subtransmission_line=%s,interruption_date=%s,extra_data=%s,form_definition_snapshot=%s,updated_at=NOW()
                WHERE id=%s RETURNING *
                """,
                (number, code, payload["company"], payload["management"], payload["associated_products"], payload["subtype"], payload["subject"], payload["observations"], payload["case_type"], payload["feeder_name"], payload["substation"], payload["subtransmission_line"], payload["interruption_date"], Json(merged_extra), Json(form_snapshot), doc_id),
            )
            item = cur.fetchone()
            cur.execute("INSERT INTO audit_log(actor_user_id,action,document_id,detail) VALUES(%s,'DOCUMENT_UPDATE',%s,jsonb_build_object('type',%s))", (actor, doc_id, doc_type))
            db.commit()
        notify_document("document_updated", item)
        return jsonify(json_ready(item))
    except (ValueError, TypeError) as exc:
        return jsonify(error=str(exc)), 400
    except psycopg2.errors.UniqueViolation:
        return jsonify(error="El correlativo o código ya existe"), 409
    except psycopg2.Error:
        app.logger.exception("Error de base de datos al actualizar el documento")
        return jsonify(error="No se pudo actualizar el documento"), 500
    except Exception:
        app.logger.exception("Error inesperado al actualizar el documento")
        return jsonify(error="No se pudo actualizar el documento"), 500


@app.delete("/api/documents/<doc_type>/<int:doc_id>")
def delete_document(doc_type, doc_id):
    denied = require_internal(admin=True)
    if denied:
        return denied
    if doc_type not in ALLOWED_TYPES:
        return jsonify(error="Tipo de documento inválido"), 400
    actor = current_user_id()
    with conn() as db, db.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute("DELETE FROM documents WHERE id=%s AND document_type=%s RETURNING *", (doc_id, doc_type))
        item = cur.fetchone()
        if not item:
            return jsonify(error="Documento no encontrado"), 404
        cur.execute("INSERT INTO audit_log(actor_user_id,action,document_id,detail) VALUES(%s,'DOCUMENT_DELETE',%s,jsonb_build_object('type',%s))", (actor, doc_id, doc_type))
        db.commit()
    notify_document("document_deleted", item)
    return jsonify(ok=True)


def excel_safe(value):
    if value is None:
        return ""
    text = str(value)
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


@app.get("/api/documents/<doc_type>/export.xlsx")
def export_documents(doc_type):
    denied = require_internal()
    if denied:
        return denied
    if doc_type not in ALLOWED_TYPES:
        return jsonify(error="Tipo de documento inválido"), 400
    user_id = request.args.get("user_id", type=int)
    if request.headers.get("X-User-Role") != "admin" and current_user_id() <= 0:
        return jsonify(error="Usuario inválido"), 400
    conditions = ["document_type=%s"]
    params = [doc_type]
    if user_id:
        conditions.append("user_id=%s")
        params.append(user_id)
    with conn() as db, db.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(f"SELECT * FROM documents WHERE {' AND '.join(conditions)} ORDER BY year DESC,number DESC", params)
        rows = cur.fetchall()
    columns = [
        ("Código", "code"), ("Empresa", "company"), ("Gestión", "management"),
        ("Productos asociados", "associated_products"), ("Tipo", "subtype"),
        ("Caso fortuito", "case_type"), ("Alimentador", "feeder_name"),
        ("Subestación", "substation"), ("Línea de subtransmisión", "subtransmission_line"),
        ("Fecha de interrupción", "interruption_date"), ("Asunto", "subject"),
        ("Observaciones", "observations"), ("Fecha", "document_date"),
        ("Hora", "document_time"), ("ID funcionario", "user_id"),
    ]
    if doc_type not in {"informes", "reportes"}:
        columns = [c for c in columns if c[1] not in {"subtype", "case_type", "feeder_name", "substation", "subtransmission_line", "interruption_date"}]
    elif doc_type == "reportes":
        columns = [c for c in columns if c[1] not in {"case_type", "feeder_name", "substation", "subtransmission_line", "interruption_date"}]
    custom_keys = []
    for row in rows:
        for key in (row.get("extra_data") or {}).keys():
            if key not in custom_keys:
                custom_keys.append(key)
    custom_keys.sort()
    columns.extend([("Campo adicional: " + key.replace("_", " ").title(), "extra:" + key) for key in custom_keys])
    wb = Workbook()
    ws = wb.active
    ws.title = doc_type.capitalize()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for col_idx, (label, _) in enumerate(columns, 1):
        cell = ws.cell(1, col_idx, label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, (_, key) in enumerate(columns, 1):
            value = (row.get("extra_data") or {}).get(key[6:]) if key.startswith("extra:") else row.get(key)
            cell = ws.cell(row_idx, col_idx, excel_safe(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for idx, (label, key) in enumerate(columns, 1):
        values = [((r.get("extra_data") or {}).get(key[6:]) if key.startswith("extra:") else r.get(key)) for r in rows]
        max_len = max([len(label)] + [len(excel_safe(value)) for value in values])
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 55)
    ws.freeze_panes = "A2"
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(stream, as_attachment=True, download_name=f"{doc_type}_{datetime.now().date().isoformat()}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=8000)
