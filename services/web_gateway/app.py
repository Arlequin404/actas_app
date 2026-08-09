import os
import secrets
import io
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from functools import wraps
from urllib.parse import quote

import requests
from flask import Flask, Response, flash, jsonify, redirect, render_template, request, session, url_for, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SERVICE_DIR = Path(__file__).resolve().parent
app = Flask(
    __name__,
    template_folder=str(SERVICE_DIR / "templates"),
    static_folder=str(SERVICE_DIR / "static"),
)
secret = os.getenv("SECRET_KEY", "")
if len(secret) < 32 or secret == "por_favor_cambia_esta_clave":
    raise RuntimeError("SECRET_KEY debe ser una cadena aleatoria de al menos 32 caracteres")
app.secret_key = secret
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=os.getenv("SESSION_COOKIE_SECURE", "false").lower() == "true",
    MAX_CONTENT_LENGTH=int(os.getenv("MAX_UPLOAD_BYTES", str(100 * 1024 * 1024))),
)
INTERNAL_API_KEY = os.environ["INTERNAL_API_KEY"]
if len(INTERNAL_API_KEY) < 32:
    raise RuntimeError("INTERNAL_API_KEY debe tener al menos 32 caracteres")
AUTH_URL = os.getenv("AUTH_SERVICE_URL", "http://auth-service:8000")
DOCUMENT_URL = os.getenv("DOCUMENT_SERVICE_URL", "http://document-service:8000")
CATALOG_URL = os.getenv("CATALOG_SERVICE_URL", "http://catalog-service:8000")
BACKUP_URL = os.getenv("BACKUP_SERVICE_URL", "http://backup-service:8000")
NOTIFICATION_URL = os.getenv("NOTIFICATION_SERVICE_URL", "http://notification-service:8000")
APP_BASE_URL = os.getenv("APP_BASE_URL", "http://localhost:8080").rstrip("/")
TIMEOUT = (4, 30)
ALLOWED_TYPES = {"actas", "informes", "reportes", "comisiones"}
DOC_TYPE_LABELS = {"actas": "Actas", "informes": "Informes", "reportes": "Reportes", "comisiones": "Comisiones"}
FIELD_TYPE_LABELS = {
    "text": "Texto corto", "textarea": "Texto largo", "number": "Número",
    "date": "Fecha", "datetime": "Fecha y hora", "time": "Hora",
    "email": "Correo electrónico", "tel": "Teléfono", "url": "Enlace web",
    "select": "Lista desplegable", "radio": "Opción única",
    "multiselect": "Selección múltiple (lista)", "checkboxes": "Selección múltiple (casillas)",
    "checkbox": "Casilla de confirmación", "yesno": "Sí / No",
}


@app.after_request
def security_headers(response):
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "DENY")
    response.headers.setdefault("Referrer-Policy", "same-origin")
    response.headers.setdefault("Permissions-Policy", "camera=(), microphone=(), geolocation=()")
    response.headers.setdefault(
        "Content-Security-Policy",
        "default-src 'self'; style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
        "font-src 'self' https://cdn.jsdelivr.net data:; img-src 'self' data:; connect-src 'self'",
    )
    if app.config["SESSION_COOKIE_SECURE"]:
        response.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
    return response


def csrf_token():
    if "csrf_token" not in session:
        session["csrf_token"] = secrets.token_urlsafe(32)
    return session["csrf_token"]


@app.context_processor
def inject_security():
    return {"csrf_token": csrf_token}


@app.before_request
def verify_csrf():
    if request.method in {"POST", "PUT", "PATCH", "DELETE"}:
        provided = request.form.get("csrf_token") or request.headers.get("X-CSRF-Token")
        expected = session.get("csrf_token")
        if not expected or not provided or not secrets.compare_digest(expected, provided):
            return "Solicitud rechazada por protección CSRF", 400


def service_headers():
    headers = {"X-Internal-Key": INTERNAL_API_KEY, "X-Client-IP": request.remote_addr or "unknown"}
    if session.get("user_id"):
        headers["X-User-ID"] = str(session["user_id"])
        headers["X-User-Role"] = str(session.get("rol", ""))
    return headers


def api(method, base, path, **kwargs):
    headers = service_headers()
    headers.update(kwargs.pop("headers", {}))
    try:
        return requests.request(method, f"{base}{path}", headers=headers, timeout=TIMEOUT, **kwargs)
    except requests.RequestException:
        app.logger.exception("Servicio no disponible: %s%s", base, path)
        return None


def current_user_profile():
    user_id = session.get("user_id")
    if not user_id:
        return None
    resp = api("GET", AUTH_URL, f"/api/users/{user_id}")
    return resp.json() if resp and resp.status_code == 200 else None


def send_system_notification(event_type, recipient, context):
    if not recipient:
        return
    api(
        "POST",
        NOTIFICATION_URL,
        "/api/notifications",
        json={"event_type": event_type, "recipient": recipient, "context": context},
    )


def write_system_log(event_type, module, level="info", detail=None, actor_name=None, actor_user_id=None, actor_role=None):
    payload = {
        "event_type": event_type,
        "module": module,
        "level": level,
        "actor_user_id": actor_user_id if actor_user_id is not None else session.get("user_id"),
        "actor_name": actor_name if actor_name is not None else session.get("nombre"),
        "actor_role": actor_role if actor_role is not None else session.get("rol"),
        "ip_address": request.headers.get("X-Forwarded-For", request.remote_addr or "unknown").split(",")[0].strip(),
        "detail": detail or {},
    }
    try:
        api("POST", NOTIFICATION_URL, "/api/system-logs", json=payload)
    except Exception:
        app.logger.exception("No fue posible registrar el log del sistema")


def response_error(resp, fallback="No fue posible completar la operación"):
    if resp is None:
        return fallback
    try:
        return resp.json().get("error") or resp.json().get("message") or fallback
    except ValueError:
        return fallback


def refresh_user():
    uid = session.get("user_id")
    if not uid:
        return False
    resp = api("GET", AUTH_URL, f"/api/users/{uid}")
    if not resp or resp.status_code != 200:
        session.clear()
        return False
    user = resp.json()
    if not user.get("activo") or user.get("session_version") != session.get("session_version"):
        session.clear()
        return False
    session["nombre"] = user["nombre"]
    session["rol"] = user["rol"]
    return True


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not refresh_user():
            flash("Inicie sesión para continuar.", "warning")
            return redirect(url_for("index"))
        return view(*args, **kwargs)
    return wrapped


def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not refresh_user():
            flash("Inicie sesión para continuar.", "warning")
            return redirect(url_for("index"))
        if session.get("rol") != "admin":
            flash("Acceso denegado.", "danger")
            return redirect(url_for("dashboard"))
        return view(*args, **kwargs)
    return wrapped


def catalog(category, include_inactive=False):
    suffix = "?include_inactive=1" if include_inactive else ""
    resp = api("GET", CATALOG_URL, f"/api/catalogs/{category}{suffix}")
    items = resp.json().get("items", []) if resp and resp.status_code == 200 else []
    # Las listas fijas del formulario reconocen el valor "Otros" para mostrar
    # la caja de escritura manual. El administrador puede configurar ese
    # comportamiento sin depender del texto visible de la opción.
    for item in items:
        if (item.get("meta_data") or {}).get("special") == "OTROS":
            item["valor"] = "Otros"
    return items


def report_tree():
    resp = api("GET", CATALOG_URL, "/api/catalogs/TIPO_INFORME/tree")
    return resp.json().get("items", []) if resp and resp.status_code == 200 else []


def form_fields(doc_type, include_inactive=False):
    suffix = "?include_inactive=1" if include_inactive else ""
    resp = api("GET", CATALOG_URL, f"/api/form-fields/{doc_type}{suffix}")
    return resp.json().get("items", []) if resp and resp.status_code == 200 else []


def form_sections(doc_type, include_inactive=False):
    suffix = "?include_inactive=1" if include_inactive else ""
    resp = api("GET", CATALOG_URL, f"/api/form-sections/{doc_type}{suffix}")
    return resp.json().get("items", []) if resp and resp.status_code == 200 else []


def form_shortcuts(include_inactive=False):
    suffix = "?include_inactive=1" if include_inactive else ""
    resp = api("GET", CATALOG_URL, f"/api/form-shortcuts{suffix}")
    return resp.json().get("items", []) if resp and resp.status_code == 200 else []


def form_setting(setting_key, default=None):
    resp = api("GET", CATALOG_URL, f"/api/settings/{setting_key}")
    if resp and resp.status_code == 200:
        return resp.json().get("setting_value", default or {})
    return default or {}


def condition_value(field_key):
    if field_key.startswith("custom:"):
        key = field_key.split(":", 1)[1]
        values = request.form.getlist(f"custom_{key}")
        return values if len(values) > 1 else (values[0] if values else "")
    mapping = {
        "tipo_informe": "tipo_informe",
        "tipo_reporte": "tipo_reporte",
        "empresa": "empresa",
        "gestiones": "gestiones_reporte" if request.view_args and request.view_args.get("doc_type") == "reportes" else "gestiones",
        "productos_asociados": "productos_asociados_reporte" if request.view_args and request.view_args.get("doc_type") == "reportes" else "productos_asociados",
    }
    return request.form.get(mapping.get(field_key, field_key), "")


def condition_matches(condition):
    if not condition:
        return True
    actual = condition_value(str(condition.get("field_key", "")))
    expected = str(condition.get("value", "")).strip()
    operator = str(condition.get("operator", "equals")).lower()
    actual_values = actual if isinstance(actual, list) else [actual]
    if operator == "not_empty":
        return any(str(value).strip() for value in actual_values)
    if operator == "contains":
        return any(expected.lower() in str(value).lower() for value in actual_values)
    return any(str(value).strip().lower() == expected.lower() for value in actual_values)


def custom_form_values(doc_type):
    values = {}
    sections = {section["id"]: section for section in form_sections(doc_type)}
    for field in form_fields(doc_type):
        section = sections.get(field.get("section_id"))
        if section and not condition_matches(section.get("show_when") or {}):
            continue
        key = field["field_key"]
        form_name = f"custom_{key}"
        field_type = field["field_type"]
        allow_other = bool(field.get("allow_other"))
        options = field.get("options") or []

        if field_type == "checkbox":
            raw = request.form.get(form_name, "")
            value = raw == "1"
        elif field_type == "yesno":
            raw = request.form.get(form_name, "")
            value = True if raw == "1" else (False if raw == "0" else "")
        elif field_type in {"multiselect", "checkboxes"}:
            selected = [item.strip() for item in request.form.getlist(form_name) if item.strip()]
            other_selected = "__other__" in selected
            value = ["Otros" if item == "__other__" else item for item in selected]
            invalid = [item for item in value if item != "Otros" and item not in options]
            if invalid:
                raise ValueError(f"La selección en '{field['label']}' no es válida")
            if other_selected:
                other_text = request.form.get(f"{form_name}__other", "").strip()
                if not allow_other or not other_text:
                    raise ValueError(f"Debe especificar la opción Otros en '{field['label']}'")
                values[f"{key}__other"] = other_text
        else:
            value = request.form.get(form_name, "").strip()
            if field_type in {"select", "radio"}:
                if value == "__other__":
                    other_text = request.form.get(f"{form_name}__other", "").strip()
                    if not allow_other or not other_text:
                        raise ValueError(f"Debe especificar la opción Otros en '{field['label']}'")
                    value = "Otros"
                    values[f"{key}__other"] = other_text
                elif value and value not in options:
                    raise ValueError(f"La opción seleccionada en '{field['label']}' no es válida")

        empty = value == "" or value == [] or (field_type == "checkbox" and value is False)
        if field.get("required") and empty:
            raise ValueError(f"El campo '{field['label']}' es obligatorio")
        if field_type == "number" and value != "":
            try:
                value = float(value)
                if value.is_integer():
                    value = int(value)
            except ValueError as exc:
                raise ValueError(f"El campo '{field['label']}' debe ser numérico") from exc
        if isinstance(value, str) and len(value) > 5000:
            raise ValueError(f"El campo '{field['label']}' es demasiado extenso")
        values[key] = value
    return values



def category_for(doc_type, field):
    if field == "management":
        return {"actas": "GESTION_ACTA", "informes": "GESTION_INFORME", "reportes": "GESTION_REPORTE", "comisiones": "GESTION_COMISION"}[doc_type]
    return {"actas": "PRODUCTO_ACTA", "informes": "PRODUCTO_INFORME", "reportes": "PRODUCTO_REPORTE", "comisiones": "PRODUCTO_COMISION"}[doc_type]


def all_companies():
    values = {}
    for item in catalog("EMPRESA"):
        value = item.get("valor") or item["nombre"]
        values[value] = item["nombre"]
    resp = api("GET", DOCUMENT_URL, "/api/documents/companies")
    if resp and resp.status_code == 200:
        for value in resp.json().get("items", []):
            values.setdefault(value, value)
    return [{"nombre": values[value], "valor": value} for value in sorted(values, key=str.lower)]


def resolve_other(value, options):
    known = {item.get("valor") or item.get("nombre") for item in options}
    return (value, "") if value in known else ("Otros", value or "")


def catalog_value_is_other(value, options):
    normalized = str(value or "").strip().lower()
    if normalized == "otros":
        return True
    for item in options:
        item_value = str(item.get("valor") or item.get("nombre") or "").strip()
        special = str((item.get("meta_data") or {}).get("special") or "").upper()
        if item_value == value and special == "OTROS":
            return True
    return False


def form_payload(doc_type):
    company = request.form.get("empresa", "").strip()
    if company == "Otros":
        company = request.form.get("empresa_otro", "").strip()

    management_options = catalog(category_for(doc_type, "management"))
    product_options = catalog(category_for(doc_type, "product"))
    if doc_type == "reportes":
        management = request.form.get("gestiones_reporte", "").strip()
        if catalog_value_is_other(management, management_options):
            management = request.form.get("gestiones_reporte_otro", "").strip()
        product = request.form.get("productos_asociados_reporte", "").strip()
        if catalog_value_is_other(product, product_options):
            product = request.form.get("productos_asociados_reporte_otro", "").strip()
        report_options = catalog("TIPO_REPORTE")
        subtype = request.form.get("tipo_reporte", "").strip()
        if catalog_value_is_other(subtype, report_options):
            subtype = request.form.get("tipo_reporte_otro", "").strip()
        type_key = "tipo_reporte"
    else:
        management = request.form.get("gestiones", "").strip()
        if catalog_value_is_other(management, management_options):
            management = request.form.get("gestiones_otro", "").strip()
        product = request.form.get("productos_asociados", "").strip()
        if catalog_value_is_other(product, product_options):
            product = request.form.get("productos_asociados_otro", "").strip()
        subtype = request.form.get("tipo_informe", "").strip() if doc_type == "informes" else ""
        if doc_type == "informes":
            # El árbol de tipos puede marcar cualquier opción como OTROS.
            all_report_nodes = catalog("TIPO_INFORME")
            if catalog_value_is_other(subtype, all_report_nodes):
                subtype = request.form.get("tipo_informe_otro", "").strip()
        type_key = "tipo_informe"
    custom_fields = custom_form_values(doc_type)
    legacy_case_type = str(custom_fields.get("elemento_afectado") or request.form.get("caso_tipo", "")).strip()
    if legacy_case_type not in {"ALIMENTADOR", "LINEAS DE SUBTRANSMISION"}:
        legacy_case_type = ""
    data = {
        "empresa": company,
        "gestiones": management,
        "productos_asociados": product,
        "asunto": request.form.get("asunto", "").strip(),
        "observaciones": request.form.get("observaciones", "").strip(),
        "numero_manual": request.form.get("numero_manual", "").strip() or None,
        type_key: subtype,
        "caso_tipo": legacy_case_type,
        "nombre_alimentador": str(custom_fields.get("nombre_alimentador") or request.form.get("nombre_alimentador", "")).strip(),
        "alimentador_subestacion": str(custom_fields.get("alimentador_subestacion") or request.form.get("alimentador_subestacion", "")).strip(),
        "linea_subtransmision_nombre": str(custom_fields.get("linea_subtransmision_nombre") or request.form.get("linea_subtransmision_nombre", "")).strip(),
        "fecha_interrupcion": str(custom_fields.get("fecha_interrupcion") or request.form.get("fecha_interrupcion", "")).strip(),
        "custom_fields": custom_fields,
        "form_definition_snapshot": form_fields(doc_type),
    }
    return data


def form_context(doc_type, doc=None, presets=None):
    presets = presets or {}
    companies = all_companies()
    managements = catalog(category_for(doc_type, "management"))
    products = catalog(category_for(doc_type, "product"))
    report_types = catalog("TIPO_REPORTE") if doc_type == "reportes" else []
    active_custom_fields = form_fields(doc_type)
    active_sections = form_sections(doc_type)
    custom_values = dict((doc or {}).get("extra_data") or {})
    if doc and doc_type == "informes":
        custom_values.setdefault("elemento_afectado", doc.get("case_type") or "")
        custom_values.setdefault("fecha_interrupcion", str(doc.get("interruption_date") or ""))
        custom_values.setdefault("nombre_alimentador", doc.get("feeder_name") or "")
        custom_values.setdefault("alimentador_subestacion", doc.get("substation") or "")
        custom_values.setdefault("linea_subtransmision_nombre", doc.get("subtransmission_line") or "")
    archived_custom_fields = []
    if doc and custom_values:
        all_custom_fields = form_fields(doc_type, include_inactive=True)
        active_keys = {field["field_key"] for field in active_custom_fields}
        active_keys.update(f"{field['field_key']}__other" for field in active_custom_fields if field.get("allow_other"))
        definitions = {field["field_key"]: field for field in all_custom_fields}
        for key, value in custom_values.items():
            if key.endswith("__other"):
                continue
            if key not in active_keys:
                definition = definitions.get(key, {})
                other = custom_values.get(f"{key}__other")
                display = f"{value}: {other}" if other else value
                archived_custom_fields.append({
                    "field_key": key,
                    "label": definition.get("label") or key.replace("_", " ").title(),
                    "value": display,
                })

    section_map = {section["id"]: dict(section, fields=[]) for section in active_sections}
    ungrouped_fields = []
    for field in active_custom_fields:
        section = section_map.get(field.get("section_id"))
        if section:
            section["fields"].append(field)
        else:
            ungrouped_fields.append(field)
    dynamic_sections = list(section_map.values())
    if ungrouped_fields:
        dynamic_sections.append({
            "id": None,
            "section_key": "campos_adicionales",
            "title": "Información adicional",
            "description": "Preguntas adicionales configuradas para este tipo de documento.",
            "icon": "bi-ui-checks-grid",
            "show_when": {},
            "settings": {},
            "fields": ungrouped_fields,
            "section_order": 999,
        })
    dynamic_sections.sort(key=lambda item: (int(item.get("section_order") or 0), int(item.get("id") or 999999)))

    company_other = form_setting("company_other", {"enabled": True, "label": "Otros", "prompt": "Especifique la empresa"})
    ctx = {
        "tipo": doc_type,
        "empresas": companies,
        "company_other": company_other,
        "gestiones": managements,
        "productos_asociados": products,
        "tipos_reporte": report_types,
        "report_types_json": report_tree() if doc_type == "informes" else [],
        "custom_fields": active_custom_fields,
        "dynamic_sections": dynamic_sections,
        "custom_values": custom_values,
        "archived_custom_fields": archived_custom_fields,
        "editar": bool(doc),
        "causas_caso_fortuito": [],
        "document_tabs": DOC_TYPE_LABELS,
        "preset_values": presets,
    }
    if not doc:
        if doc_type == "informes" and presets.get("tipo_informe"):
            ctx["tipo_informe_sel"] = presets["tipo_informe"]
            ctx["tipo_informe"] = presets["tipo_informe"]
        if doc_type == "reportes" and presets.get("tipo_reporte"):
            ctx["tipo_reporte_sel"] = presets["tipo_reporte"]
        return ctx
    company_sel, company_other = resolve_other(doc.get("company"), companies)
    management_sel, management_other = resolve_other(doc.get("management"), managements)
    product_sel, product_other = resolve_other(doc.get("associated_products"), products)
    ctx.update({
        "id": doc["id"], "numero_original": doc["number"],
        "empresa_sel": company_sel, "empresa_otro": company_other,
        "gestiones_sel": management_sel, "gestiones_otro": management_other,
        "productos_asociados_sel": product_sel, "productos_asociados_otro": product_other,
        "asunto": doc.get("subject"), "observaciones": doc.get("observations"),
        "caso_tipo": doc.get("case_type"), "nombre_alimentador": doc.get("feeder_name"),
        "alimentador_subestacion": doc.get("substation"),
        "linea_subtransmision_nombre": doc.get("subtransmission_line"),
        "fecha_interrupcion": str(doc.get("interruption_date") or ""),
    })
    if doc_type == "informes":
        ctx["tipo_informe_sel"] = doc.get("subtype") or ""
        ctx["tipo_informe"] = doc.get("subtype") or ""
        ctx["tipo_informe_otro"] = doc.get("subtype") or ""
    elif doc_type == "reportes":
        subtype_sel, subtype_other = resolve_other(doc.get("subtype"), report_types)
        ctx["tipo_reporte_sel"] = subtype_sel
        ctx["tipo_reporte_otro"] = subtype_other
    return ctx


def format_docs(items, doc_type, user_names):
    rows = []
    for d in items:
        owner_id = d.get("user_id")
        if owner_id in user_names:
            funcionario = user_names[owner_id]
        elif owner_id == session.get("user_id"):
            funcionario = session.get("nombre", "—")
        else:
            funcionario = f"Usuario eliminado (ID {owner_id})"
        if doc_type == "informes":
            rows.append((d["id"], d["code"], d["company"], d["management"], d["associated_products"], d.get("subtype"), d.get("case_type"), d.get("feeder_name"), d.get("substation"), d.get("subtransmission_line"), d.get("interruption_date"), d["subject"], d.get("observations"), d["document_date"], d["document_time"], funcionario, owner_id))
        elif doc_type == "reportes":
            rows.append((d["id"], d["code"], d["company"], d["management"], d["associated_products"], d.get("subtype"), d["subject"], d.get("observations"), d["document_date"], d["document_time"], funcionario, owner_id))
        else:
            rows.append((d["id"], d["code"], d["company"], d["management"], d["associated_products"], d["subject"], d.get("observations"), d["document_date"], d["document_time"], funcionario, owner_id))
    return rows


def list_page(user_id=None):
    page = max(request.args.get("page", 1, type=int), 1)
    per_page = request.args.get("per_page", "10")
    company = request.args.get("empresa", "").strip()
    active_tab = request.args.get("tab", "actas")
    datasets = {}
    pages = {}
    all_user_ids = set()
    for doc_type in ALLOWED_TYPES:
        params = {"page": page, "per_page": per_page}
        if company:
            params["company"] = company
        if user_id:
            params["user_id"] = user_id
        resp = api("GET", DOCUMENT_URL, f"/api/documents/{doc_type}", params=params)
        data = resp.json() if resp and resp.status_code == 200 else {"items": [], "pages": 1}
        datasets[doc_type] = data.get("items", [])
        pages[doc_type] = data.get("pages", 1)
        all_user_ids.update(x.get("user_id") for x in datasets[doc_type])
    names = {}
    valid_user_ids = sorted(user_id for user_id in all_user_ids if isinstance(user_id, int))
    if valid_user_ids:
        resp = api("GET", AUTH_URL, "/api/users/display-names", params={"ids": ",".join(map(str, valid_user_ids))})
        if resp and resp.status_code == 200:
            names = {u["id"]: u["nombre"] for u in resp.json().get("items", [])}
    return {
        "actas": format_docs(datasets["actas"], "actas", names),
        "informes": format_docs(datasets["informes"], "informes", names),
        "reportes": format_docs(datasets["reportes"], "reportes", names),
        "comisiones": format_docs(datasets["comisiones"], "comisiones", names),
        "page": page, "per_page": per_page,
        "pages_a": pages["actas"], "pages_i": pages["informes"],
        "pages_r": pages["reportes"], "pages_c": pages["comisiones"],
        "empresas": all_companies(), "empresa_sel": company, "active_tab": active_tab,
    }


def build_admin_dashboard_data():
    """Construye un resumen global. Solo se usa desde rutas protegidas de administrador."""
    users_resp = api("GET", AUTH_URL, "/api/users")
    users = users_resp.json().get("items", []) if users_resp and users_resp.status_code == 200 else []
    user_names = {u.get("id"): u.get("nombre") or u.get("email") or f"Usuario {u.get('id')}" for u in users}

    documents_by_type = {}
    all_documents = []
    for doc_type in sorted(ALLOWED_TYPES):
        resp = api("GET", DOCUMENT_URL, f"/api/documents/{doc_type}", params={"page": 1, "per_page": "all"})
        data = resp.json() if resp and resp.status_code == 200 else {"items": [], "total": 0}
        items = data.get("items", [])
        documents_by_type[doc_type] = len(items)
        for item in items:
            row = dict(item)
            row["document_type"] = doc_type
            row["funcionario"] = user_names.get(item.get("user_id"), f"Usuario ID {item.get('user_id')}")
            all_documents.append(row)

    current_year = datetime.now().year
    current_year_count = sum(1 for d in all_documents if d.get("year") == current_year)
    company_counts = Counter((d.get("company") or "Sin empresa").strip() or "Sin empresa" for d in all_documents)
    user_doc_counts = Counter(d.get("user_id") for d in all_documents)

    by_user = []
    for user in users:
        uid = user.get("id")
        by_user.append({
            "id": uid,
            "nombre": user.get("nombre") or user.get("email") or f"Usuario {uid}",
            "email": user.get("email", ""),
            "rol": user.get("rol", ""),
            "activo": bool(user.get("activo")),
            "documentos": user_doc_counts.get(uid, 0),
        })
    # También conserva documentos de usuarios que ya no existan.
    known_ids = {u.get("id") for u in users}
    for uid, count in user_doc_counts.items():
        if uid not in known_ids:
            by_user.append({"id": uid, "nombre": f"Usuario eliminado (ID {uid})", "email": "", "rol": "—", "activo": False, "documentos": count})
    by_user.sort(key=lambda x: (-x["documentos"], x["nombre"].lower()))

    def recent_key(d):
        return (str(d.get("document_date") or ""), str(d.get("document_time") or ""), int(d.get("id") or 0))
    recent_documents = sorted(all_documents, key=recent_key, reverse=True)[:10]

    notifications_resp = api("GET", NOTIFICATION_URL, "/api/notifications", params={"page": 1, "per_page": 1})
    notification_total = notifications_resp.json().get("total", 0) if notifications_resp and notifications_resp.status_code == 200 else 0
    failed_resp = api("GET", NOTIFICATION_URL, "/api/notifications", params={"page": 1, "per_page": 1, "status": "failed"})
    failed_notifications = failed_resp.json().get("total", 0) if failed_resp and failed_resp.status_code == 200 else 0

    total_documents = len(all_documents)
    type_labels = DOC_TYPE_LABELS
    type_distribution = [
        {"key": key, "label": type_labels[key], "count": documents_by_type.get(key, 0),
         "percent": round((documents_by_type.get(key, 0) * 100 / total_documents), 1) if total_documents else 0}
        for key in ("actas", "informes", "reportes", "comisiones")
    ]
    top_companies = [
        {"name": name, "count": count, "percent": round(count * 100 / total_documents, 1) if total_documents else 0}
        for name, count in company_counts.most_common(8)
    ]

    return {
        "user_count": len(users),
        "active_user_count": sum(1 for u in users if u.get("activo")),
        "admin_count": sum(1 for u in users if u.get("rol") == "admin"),
        "total_documents": total_documents,
        "current_year": current_year,
        "current_year_count": current_year_count,
        "document_counts": documents_by_type,
        "type_distribution": type_distribution,
        "by_user": by_user,
        "top_companies": top_companies,
        "recent_documents": recent_documents,
        "notification_total": notification_total,
        "failed_notifications": failed_notifications,
    }


@app.get("/")
def index():
    if session.get("user_id") and refresh_user():
        return redirect(url_for("dashboard"))
    return render_template("login.html")


@app.post("/login")
def login():
    resp = api("POST", AUTH_URL, "/api/auth/login", json={"email": request.form.get("email", ""), "password": request.form.get("password", "")})
    if not resp or resp.status_code != 200:
        write_system_log("LOGIN_FAILED", "auth", level="warning", detail={"email": request.form.get("email", "")}, actor_name=request.form.get("email", ""), actor_role="anonymous")
        flash("Correo o contraseña incorrectos.", "danger")
        return redirect(url_for("index"))
    user = resp.json()
    session.clear()
    session.update(user_id=user["id"], nombre=user["nombre"], rol=user["rol"], session_version=user["session_version"])
    csrf_token()
    write_system_log("LOGIN_SUCCESS", "auth", detail={"email": user.get("email", "")}, actor_name=user.get("nombre"), actor_user_id=user.get("id"), actor_role=user.get("rol"))
    return redirect(url_for("dashboard"))


@app.post("/logout")
def logout():
    write_system_log("LOGOUT", "auth")
    session.clear()
    return redirect(url_for("index"))


@app.get("/dashboard")
@login_required
def dashboard():
    return render_template("dashboard.html", nombre=session["nombre"])


@app.get("/recuperar_contrasena")
@app.get("/recuperar_contraseña", endpoint="recuperar_contraseña")
def recuperar_contrasena():
    return render_template("recuperar_contraseña.html")


@app.post("/enviar_recuperacion")
def enviar_recuperacion():
    api("POST", AUTH_URL, "/api/auth/password-reset/request", json={"email": request.form.get("email", "")})
    flash("Si el correo está registrado, recibirá instrucciones.", "success")
    return redirect(url_for("index"))


@app.route("/restablecer_contrasena/<token>", methods=["GET", "POST"])
@app.route("/restablecer_contraseña/<token>", methods=["GET", "POST"], endpoint="restablecer_contraseña")
def restablecer_contrasena(token):
    if request.method == "GET":
        resp = api("GET", AUTH_URL, f"/api/auth/password-reset/{quote(token)}")
        if not resp or resp.status_code != 200:
            flash("El enlace es inválido o expiró.", "danger")
            return redirect(url_for("index"))
        return render_template("restablecer_contraseña.html", token=token)
    password = request.form.get("password", "")
    confirm = request.form.get("confirm_password", request.form.get("password_confirm", ""))
    if password != confirm:
        flash("Las contraseñas no coinciden.", "warning")
        return redirect(request.url)
    resp = api("POST", AUTH_URL, f"/api/auth/password-reset/{quote(token)}", json={"password": password})
    if not resp or resp.status_code != 200:
        flash(response_error(resp), "danger")
        return redirect(request.url)
    flash("Contraseña actualizada correctamente.", "success")
    return redirect(url_for("index"))


@app.get("/crear")
@login_required
def crear_selector():
    return render_template("crear_selector.html", shortcuts=form_shortcuts(), doc_type_labels=DOC_TYPE_LABELS)


@app.route("/crear/<doc_type>", methods=["GET", "POST"])
@login_required
def crear(doc_type):
    if doc_type not in ALLOWED_TYPES:
        flash("Tipo de documento inválido.", "danger")
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        try:
            payload = form_payload(doc_type)
        except ValueError as exc:
            flash(str(exc), "danger")
        else:
            resp = api("POST", DOCUMENT_URL, f"/api/documents/{doc_type}", json=payload)
            if resp and resp.status_code == 201:
                created = resp.json()
                write_system_log("DOCUMENT_CREATE", "documents", detail={"type": doc_type, "id": created.get("id"), "code": created.get("code")})
                flash(f"Documento {created['code']} creado correctamente.", "success")
                return redirect(url_for("mis_documentos" if session.get("rol") != "admin" else "admin_documentos", tab=doc_type))
            flash(response_error(resp), "danger")
    presets = {}
    shortcut_id = request.args.get("shortcut", type=int)
    if shortcut_id:
        shortcut = next((item for item in form_shortcuts() if item.get("id") == shortcut_id and item.get("document_type") == doc_type), None)
        if shortcut:
            presets = shortcut.get("preset_values") or {}
    return render_template("formulario.html", **form_context(doc_type, presets=presets))


@app.route("/admin/editar/<doc_type>/<int:doc_id>", methods=["GET", "POST"])
@login_required
def editar_documento(doc_type, doc_id):
    if doc_type not in ALLOWED_TYPES:
        flash("Tipo de documento inválido.", "danger")
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        try:
            payload = form_payload(doc_type)
        except ValueError as exc:
            flash(str(exc), "danger")
        else:
            resp = api("PUT", DOCUMENT_URL, f"/api/documents/{doc_type}/{doc_id}", json=payload)
            if resp and resp.status_code == 200:
                updated = resp.json()
                write_system_log("DOCUMENT_UPDATE", "documents", detail={"type": doc_type, "id": doc_id, "code": updated.get("code")})
                flash("Documento actualizado correctamente.", "success")
                return redirect(url_for("admin_documentos" if session.get("rol") == "admin" else "mis_documentos", tab=doc_type))
            flash(response_error(resp), "danger")
    resp = api("GET", DOCUMENT_URL, f"/api/documents/{doc_type}/{doc_id}")
    if not resp or resp.status_code != 200:
        flash(response_error(resp, "Documento no encontrado"), "danger")
        return redirect(url_for("dashboard"))
    document = resp.json()
    if session.get("rol") != "admin" and document.get("user_id") != session.get("user_id"):
        flash("Solo puede editar los documentos que usted creó.", "warning")
        return redirect(url_for("mis_documentos", tab=doc_type))
    return render_template("formulario.html", **form_context(doc_type, document))


@app.post("/eliminar/<doc_type>/<int:doc_id>")
@admin_required
def eliminar_documento(doc_type, doc_id):
    resp = api("DELETE", DOCUMENT_URL, f"/api/documents/{doc_type}/{doc_id}")
    if resp and resp.status_code == 200:
        write_system_log("DOCUMENT_DELETE", "documents", detail={"type": doc_type, "id": doc_id})
    flash("Documento eliminado correctamente." if resp and resp.status_code == 200 else response_error(resp), "success" if resp and resp.status_code == 200 else "danger")
    return redirect(url_for("admin_documentos", tab=doc_type))


@app.get("/exportar_documentos/<doc_type>")
@login_required
def exportar_documentos(doc_type):
    if doc_type not in ALLOWED_TYPES:
        flash("Tipo de documento inválido.", "danger")
        return redirect(url_for("dashboard"))
    resp = api("GET", DOCUMENT_URL, f"/api/documents/{doc_type}/export.xlsx", stream=True)
    if not resp or resp.status_code != 200:
        flash(response_error(resp, "No fue posible exportar los documentos"), "danger")
        return redirect(url_for("dashboard"))
    headers = {"Content-Disposition": resp.headers.get("Content-Disposition", f'attachment; filename="{doc_type}.xlsx"')}
    return Response(resp.iter_content(64 * 1024), content_type=resp.headers.get("Content-Type"), headers=headers)


@app.get("/mis_documentos")
@login_required
def mis_documentos():
    return render_template("mis_documentos.html", **list_page())


@app.get("/admin/documentos")
@admin_required
def admin_documentos():
    return render_template("admin_documentos.html", **list_page())


@app.get("/admin/dashboard")
@admin_required
def admin_dashboard():
    return render_template("admin_dashboard.html", **build_admin_dashboard_data())


@app.get("/admin/dashboard/descargar")
@admin_required
def descargar_admin_dashboard():
    data = build_admin_dashboard_data()
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    header_fill = PatternFill("solid", fgColor="0D6EFD")
    header_font = Font(color="FFFFFF", bold=True)

    summary = [
        ("Indicador", "Valor"),
        ("Usuarios totales", data["user_count"]),
        ("Usuarios activos", data["active_user_count"]),
        ("Administradores", data["admin_count"]),
        ("Documentos totales", data["total_documents"]),
        (f"Documentos {data['current_year']}", data["current_year_count"]),
        ("Actas", data["document_counts"].get("actas", 0)),
        ("Informes", data["document_counts"].get("informes", 0)),
        ("Reportes", data["document_counts"].get("reportes", 0)),
        ("Comisiones", data["document_counts"].get("comisiones", 0)),
        ("Notificaciones registradas", data["notification_total"]),
        ("Notificaciones fallidas", data["failed_notifications"]),
    ]
    for r, row in enumerate(summary, 1):
        for c, value in enumerate(row, 1):
            ws.cell(r, c, value)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18

    ws_users = wb.create_sheet("Documentos por usuario")
    user_headers = ["Funcionario", "Correo", "Rol", "Activo", "Documentos"]
    ws_users.append(user_headers)
    for item in data["by_user"]:
        ws_users.append([item["nombre"], item["email"], item["rol"], "Sí" if item["activo"] else "No", item["documentos"]])

    ws_companies = wb.create_sheet("Empresas")
    ws_companies.append(["Empresa", "Documentos", "Porcentaje"] )
    for item in data["top_companies"]:
        ws_companies.append([item["name"], item["count"], item["percent"] / 100])
    for cell in ws_companies[1]:
        cell.fill = header_fill; cell.font = header_font
    for row in ws_companies.iter_rows(min_row=2, min_col=3, max_col=3):
        row[0].number_format = "0.0%"

    ws_recent = wb.create_sheet("Últimos documentos")
    recent_headers = ["Código", "Tipo", "Funcionario", "Empresa", "Asunto", "Fecha", "Hora"]
    ws_recent.append(recent_headers)
    for d in data["recent_documents"]:
        ws_recent.append([d.get("code"), DOC_TYPE_LABELS.get(d.get("document_type"), d.get("document_type")), d.get("funcionario"), d.get("company"), d.get("subject"), d.get("document_date"), d.get("document_time")])

    for sheet in (ws_users, ws_recent):
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for col in range(1, sheet.max_column + 1):
            max_len = 10
            for cell in sheet.iter_cols(min_col=col, max_col=col, values_only=True):
                max_len = max(max_len, *(len(str(v)) for v in cell if v is not None))
            sheet.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 45)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        as_attachment=True,
        download_name=f"dashboard_administrativo_{datetime.now().date().isoformat()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/admin")
@admin_required
def admin_home():
    users_resp = api("GET", AUTH_URL, "/api/users")
    users = users_resp.json().get("items", []) if users_resp and users_resp.status_code == 200 else []
    document_counts = {}
    for doc_type in ALLOWED_TYPES:
        resp = api("GET", DOCUMENT_URL, f"/api/documents/{doc_type}", params={"page": 1, "per_page": 1})
        document_counts[doc_type] = resp.json().get("total", 0) if resp and resp.status_code == 200 else 0
    field_counts = {doc_type: len(form_fields(doc_type)) for doc_type in ALLOWED_TYPES}
    notifications_resp = api("GET", NOTIFICATION_URL, "/api/notifications", params={"page": 1, "per_page": 1})
    notification_total = notifications_resp.json().get("total", 0) if notifications_resp and notifications_resp.status_code == 200 else 0
    return render_template(
        "admin_home.html",
        user_count=len(users),
        admin_count=sum(1 for user in users if user.get("rol") == "admin"),
        active_user_count=sum(1 for user in users if user.get("activo")),
        document_counts=document_counts,
        total_documents=sum(document_counts.values()),
        field_counts=field_counts,
        notification_total=notification_total,
    )


@app.get("/admin/usuarios")
@admin_required
def admin_users():
    resp = api("GET", AUTH_URL, "/api/users")
    users = resp.json().get("items", []) if resp and resp.status_code == 200 else []
    return render_template("admin.html", usuarios=users)


@app.route("/admin/usuarios/crear", methods=["GET", "POST"])
@admin_required
def crear_usuario():
    if request.method == "POST":
        payload = {k: request.form.get(k, "") for k in ("nombre", "email", "password", "rol")}
        payload["activo"] = request.form.get("activo", "1") == "1"
        resp = api("POST", AUTH_URL, "/api/users", json=payload)
        if resp and resp.status_code == 201:
            created_user = resp.json()
            write_system_log("USER_CREATE", "users", detail={"id": created_user.get("id"), "email": created_user.get("email"), "role": created_user.get("rol")})
            flash("Usuario creado correctamente.", "success")
            return redirect(url_for("admin_users"))
        flash(response_error(resp), "danger")
    return render_template("form_usuario.html", modo="Crear", usuario=None)


@app.route("/admin/usuarios/editar/<int:user_id>", methods=["GET", "POST"])
@admin_required
def editar_usuario(user_id):
    if request.method == "POST":
        payload = {k: request.form.get(k, "") for k in ("nombre", "email", "password", "rol")}
        payload["activo"] = request.form.get("activo") == "1"
        resp = api("PUT", AUTH_URL, f"/api/users/{user_id}", json=payload)
        if resp and resp.status_code == 200:
            updated = resp.json()
            if user_id == session.get("user_id"):
                session.update(nombre=updated["nombre"], rol=updated["rol"], session_version=updated["session_version"])
            write_system_log("USER_UPDATE", "users", detail={"id": user_id, "email": updated.get("email"), "role": updated.get("rol"), "active": updated.get("activo")})
            flash("Usuario actualizado correctamente.", "success")
            return redirect(url_for("admin_users"))
        flash(response_error(resp), "danger")
    resp = api("GET", AUTH_URL, f"/api/users/{user_id}")
    if not resp or resp.status_code != 200:
        flash("Usuario no encontrado.", "danger")
        return redirect(url_for("admin_users"))
    u = resp.json()
    return render_template("form_usuario.html", modo="Editar", usuario=(u["id"], u["nombre"], u["email"], u["rol"], u["activo"]))


@app.post("/admin/usuarios/eliminar/<int:user_id>")
@admin_required
def eliminar_usuario(user_id):
    resp = api("DELETE", AUTH_URL, f"/api/users/{user_id}")
    if resp and resp.status_code == 200:
        write_system_log("USER_DEACTIVATE", "users", detail={"id": user_id})
    flash("Acceso desactivado; sus documentos y su historial se conservaron." if resp and resp.status_code == 200 else response_error(resp), "success" if resp and resp.status_code == 200 else "danger")
    return redirect(url_for("admin_users"))


@app.get("/admin/configuracion")
@admin_required
def admin_configuracion():
    # El editor visual reemplaza la pantalla de catálogos separada para evitar configuraciones duplicadas y confusas.
    return redirect(url_for("admin_campos"))


@app.get("/api/catalogos/<category>")
@admin_required
def api_get_catalogos(category):
    resp = api("GET", CATALOG_URL, f"/api/catalogs/{category}?include_inactive=1")
    if not resp:
        return jsonify(error="Servicio de catálogos no disponible"), 503
    data = resp.json()
    return jsonify(data=data.get("items", []), error=data.get("error")), resp.status_code


@app.post("/api/catalogos")
@admin_required
def api_create_catalogo():
    resp = api("POST", CATALOG_URL, "/api/catalogs", json=request.get_json(silent=True) or {})
    return jsonify(resp.json() if resp else {"error": "Servicio no disponible"}), (resp.status_code if resp else 503)


@app.post("/api/catalogos/lote")
@admin_required
def api_create_catalogos_lote():
    resp = api("POST", CATALOG_URL, "/api/catalogs/bulk", json=request.get_json(silent=True) or {})
    return jsonify(resp.json() if resp else {"error": "Servicio no disponible"}), (resp.status_code if resp else 503)


@app.post("/api/catalogos/reordenar")
@admin_required
def api_reorder_catalogos():
    resp = api("POST", CATALOG_URL, "/api/catalogs/reorder", json=request.get_json(silent=True) or {})
    return jsonify(resp.json() if resp else {"error": "Servicio no disponible"}), (resp.status_code if resp else 503)


@app.put("/api/catalogos/<int:item_id>")
@admin_required
def api_update_catalogo(item_id):
    resp = api("PUT", CATALOG_URL, f"/api/catalogs/{item_id}", json=request.get_json(silent=True) or {})
    return jsonify(resp.json() if resp else {"error": "Servicio no disponible"}), (resp.status_code if resp else 503)


@app.delete("/api/catalogos/<int:item_id>")
@admin_required
def api_delete_catalogo(item_id):
    resp = api("DELETE", CATALOG_URL, f"/api/catalogs/{item_id}")
    return jsonify(resp.json() if resp else {"error": "Servicio no disponible"}), (resp.status_code if resp else 503)


@app.delete("/api/catalogos/<int:item_id>/eliminar")
@admin_required
def api_purge_catalogo(item_id):
    cascade = "1" if request.args.get("cascade") == "1" else "0"
    resp = api("DELETE", CATALOG_URL, f"/api/catalogs/{item_id}/purge?cascade={cascade}")
    return jsonify(resp.json() if resp else {"error": "Servicio no disponible"}), (resp.status_code if resp else 503)


@app.get("/api/configuracion/<setting_key>")
@admin_required
def api_get_form_setting(setting_key):
    resp = api("GET", CATALOG_URL, f"/api/settings/{setting_key}")
    return jsonify(resp.json() if resp else {"error": "Servicio no disponible"}), (resp.status_code if resp else 503)


@app.put("/api/configuracion/<setting_key>")
@admin_required
def api_update_form_setting(setting_key):
    resp = api("PUT", CATALOG_URL, f"/api/settings/{setting_key}", json=request.get_json(silent=True) or {})
    return jsonify(resp.json() if resp else {"error": "Servicio no disponible"}), (resp.status_code if resp else 503)


@app.get("/admin/empresas")
@admin_required
def admin_empresas():
    return render_template("admin_empresas.html")


@app.get("/admin/campos")
@admin_required
def admin_campos():
    return render_template(
        "admin_campos.html",
        doc_type_labels=DOC_TYPE_LABELS,
        field_type_labels=FIELD_TYPE_LABELS,
    )


@app.get("/api/campos/<doc_type>")
@admin_required
def api_get_campos(doc_type):
    if doc_type not in ALLOWED_TYPES:
        return jsonify(error="Tipo de documento inválido"), 400
    resp = api("GET", CATALOG_URL, f"/api/form-fields/{doc_type}?include_inactive=1")
    return jsonify(resp.json() if resp else {"error": "Servicio no disponible"}), (resp.status_code if resp else 503)


@app.post("/api/campos")
@admin_required
def api_create_campo():
    resp = api("POST", CATALOG_URL, "/api/form-fields", json=request.get_json(silent=True) or {})
    return jsonify(resp.json() if resp else {"error": "Servicio no disponible"}), (resp.status_code if resp else 503)


@app.put("/api/campos/<int:field_id>")
@admin_required
def api_update_campo(field_id):
    resp = api("PUT", CATALOG_URL, f"/api/form-fields/{field_id}", json=request.get_json(silent=True) or {})
    return jsonify(resp.json() if resp else {"error": "Servicio no disponible"}), (resp.status_code if resp else 503)


@app.delete("/api/campos/<int:field_id>")
@admin_required
def api_archive_campo(field_id):
    resp = api("DELETE", CATALOG_URL, f"/api/form-fields/{field_id}")
    return jsonify(resp.json() if resp else {"error": "Servicio no disponible"}), (resp.status_code if resp else 503)


@app.delete("/api/campos/<int:field_id>/eliminar")
@admin_required
def api_purge_campo(field_id):
    cascade = "1" if request.args.get("cascade") == "1" else "0"
    resp = api("DELETE", CATALOG_URL, f"/api/form-fields/{field_id}/purge?cascade={cascade}")
    return jsonify(resp.json() if resp else {"error": "Servicio no disponible"}), (resp.status_code if resp else 503)


@app.post("/api/campos/reordenar")
@admin_required
def api_reorder_campos():
    resp = api("POST", CATALOG_URL, "/api/form-fields/reorder", json=request.get_json(silent=True) or {})
    return jsonify(resp.json() if resp else {"error": "Servicio no disponible"}), (resp.status_code if resp else 503)


@app.get("/api/secciones/<doc_type>")
@admin_required
def api_get_secciones(doc_type):
    if doc_type not in ALLOWED_TYPES:
        return jsonify(error="Tipo de documento inválido"), 400
    resp = api("GET", CATALOG_URL, f"/api/form-sections/{doc_type}?include_inactive=1")
    return jsonify(resp.json() if resp else {"error": "Servicio no disponible"}), (resp.status_code if resp else 503)


@app.post("/api/secciones")
@admin_required
def api_create_seccion():
    resp = api("POST", CATALOG_URL, "/api/form-sections", json=request.get_json(silent=True) or {})
    return jsonify(resp.json() if resp else {"error": "Servicio no disponible"}), (resp.status_code if resp else 503)


@app.put("/api/secciones/<int:section_id>")
@admin_required
def api_update_seccion(section_id):
    resp = api("PUT", CATALOG_URL, f"/api/form-sections/{section_id}", json=request.get_json(silent=True) or {})
    return jsonify(resp.json() if resp else {"error": "Servicio no disponible"}), (resp.status_code if resp else 503)


@app.delete("/api/secciones/<int:section_id>")
@admin_required
def api_archive_seccion(section_id):
    resp = api("DELETE", CATALOG_URL, f"/api/form-sections/{section_id}")
    return jsonify(resp.json() if resp else {"error": "Servicio no disponible"}), (resp.status_code if resp else 503)


@app.delete("/api/secciones/<int:section_id>/eliminar")
@admin_required
def api_purge_seccion(section_id):
    cascade = "1" if request.args.get("cascade") == "1" else "0"
    resp = api("DELETE", CATALOG_URL, f"/api/form-sections/{section_id}/purge?cascade={cascade}")
    return jsonify(resp.json() if resp else {"error": "Servicio no disponible"}), (resp.status_code if resp else 503)


@app.post("/api/secciones/<int:section_id>/copiar")
@admin_required
def api_clone_seccion(section_id):
    resp = api("POST", CATALOG_URL, f"/api/form-sections/{section_id}/clone", json=request.get_json(silent=True) or {})
    return jsonify(resp.json() if resp else {"error": "Servicio no disponible"}), (resp.status_code if resp else 503)


@app.get("/api/accesos-formulario")
@admin_required
def api_get_accesos_formulario():
    resp = api("GET", CATALOG_URL, "/api/form-shortcuts?include_inactive=1")
    return jsonify(resp.json() if resp else {"error": "Servicio no disponible"}), (resp.status_code if resp else 503)


@app.post("/api/accesos-formulario")
@admin_required
def api_create_acceso_formulario():
    resp = api("POST", CATALOG_URL, "/api/form-shortcuts", json=request.get_json(silent=True) or {})
    return jsonify(resp.json() if resp else {"error": "Servicio no disponible"}), (resp.status_code if resp else 503)


@app.put("/api/accesos-formulario/<int:shortcut_id>")
@admin_required
def api_update_acceso_formulario(shortcut_id):
    resp = api("PUT", CATALOG_URL, f"/api/form-shortcuts/{shortcut_id}", json=request.get_json(silent=True) or {})
    return jsonify(resp.json() if resp else {"error": "Servicio no disponible"}), (resp.status_code if resp else 503)


@app.delete("/api/accesos-formulario/<int:shortcut_id>")
@admin_required
def api_archive_acceso_formulario(shortcut_id):
    resp = api("DELETE", CATALOG_URL, f"/api/form-shortcuts/{shortcut_id}")
    return jsonify(resp.json() if resp else {"error": "Servicio no disponible"}), (resp.status_code if resp else 503)


@app.get("/admin/notificaciones")
@admin_required
def admin_notificaciones():
    page = max(request.args.get("page", 1, type=int), 1)
    status = request.args.get("status", "").strip()
    recipient = request.args.get("recipient", "").strip()
    resp = api("GET", NOTIFICATION_URL, "/api/notifications", params={"page": page, "per_page": 25, "status": status, "recipient": recipient})
    data = resp.json() if resp and resp.status_code == 200 else {"items": [], "total": 0}
    return render_template("admin_notificaciones.html", notifications=data.get("items", []), total=data.get("total", 0), page=page, status=status, recipient=recipient)


@app.post("/admin/notificaciones/<int:notification_id>/reintentar")
@admin_required
def reintentar_notificacion(notification_id):
    resp = api("POST", NOTIFICATION_URL, f"/api/notifications/{notification_id}/retry")
    if resp and resp.status_code == 200:
        flash("Notificación procesada nuevamente.", "success")
    else:
        flash(response_error(resp, "No fue posible reintentar la notificación"), "danger")
    return redirect(url_for("admin_notificaciones"))


@app.get("/admin/respaldo")
@admin_required
def admin_respaldo():
    resp = api("GET", BACKUP_URL, "/api/backups/status")
    backup_status = resp.json() if resp and resp.status_code == 200 else {"status": {}, "auto_enabled": False}
    return render_template("admin_respaldo.html", backup_status=backup_status)


@app.get("/admin/respaldo/google/conectar")
@admin_required
def conectar_google_drive():
    redirect_uri = APP_BASE_URL + url_for("google_drive_callback")
    resp = api(
        "POST",
        BACKUP_URL,
        "/api/google-drive/auth-url",
        json={"redirect_uri": redirect_uri},
    )
    if not resp or resp.status_code != 200:
        flash(response_error(resp, "No fue posible iniciar la conexión con Google Drive"), "danger")
        return redirect(url_for("admin_respaldo"))
    auth_url = resp.json().get("auth_url")
    if not auth_url:
        flash("Google Drive no devolvió una URL de autorización.", "danger")
        return redirect(url_for("admin_respaldo"))
    return redirect(auth_url)


@app.get("/admin/respaldo/google/callback")
@admin_required
def google_drive_callback():
    if request.args.get("error"):
        flash(f"Google Drive rechazó la autorización: {request.args.get('error')}", "warning")
        return redirect(url_for("admin_respaldo"))
    code = request.args.get("code", "").strip()
    state = request.args.get("state", "").strip()
    if not code or not state:
        flash("La respuesta de Google Drive está incompleta.", "danger")
        return redirect(url_for("admin_respaldo"))
    redirect_uri = APP_BASE_URL + url_for("google_drive_callback")
    resp = api(
        "POST",
        BACKUP_URL,
        "/api/google-drive/oauth-callback",
        json={"code": code, "state": state, "redirect_uri": redirect_uri},
    )
    if resp and resp.status_code == 200:
        account = (resp.json().get("account") or {})
        label = account.get("email") or account.get("display_name") or "la cuenta seleccionada"
        flash(f"Google Drive conectado correctamente con {label}.", "success")
    else:
        flash(response_error(resp, "No fue posible completar la conexión con Google Drive"), "danger")
    return redirect(url_for("admin_respaldo"))


@app.post("/admin/respaldo/google/desconectar")
@admin_required
def desconectar_google_drive():
    resp = api("POST", BACKUP_URL, "/api/google-drive/disconnect", json={})
    if resp and resp.status_code == 200:
        flash("Google Drive fue desconectado. Los respaldos ya existentes permanecen en la nube.", "success")
    else:
        flash(response_error(resp, "No fue posible desconectar Google Drive"), "danger")
    return redirect(url_for("admin_respaldo"))


@app.post("/admin/respaldo/google-drive-ahora")
@app.post("/admin/respaldo/onedrive-ahora")  # compatibilidad con marcadores antiguos
@admin_required
def respaldo_google_drive_ahora():
    profile = current_user_profile() or {}
    resp = api("POST", BACKUP_URL, "/api/backups/google-drive", json={"actor_name": profile.get("nombre") or session.get("nombre"), "actor_user_id": session.get("user_id")})
    if resp and resp.status_code == 200:
        flash("Respaldo subido directamente a Google Drive.", "success")
    else:
        flash(response_error(resp, "No fue posible guardar el respaldo en Google Drive"), "danger")
    return redirect(url_for("admin_respaldo"))


@app.get("/admin/respaldo/exportar")
@admin_required
def exportar_db():
    resp = api("GET", BACKUP_URL, "/api/backups/export", stream=True)
    if not resp or resp.status_code != 200:
        flash(response_error(resp, "No fue posible generar el respaldo"), "danger")
        return redirect(url_for("admin_respaldo"))
    profile = current_user_profile()
    write_system_log("BACKUP_EXPORT", "backup", detail={"databases": ["auth_db", "documents_db", "catalog_db", "notifications_db"]})
    send_system_notification(
        "backup_exported",
        (profile or {}).get("email"),
        {"actor_name": (profile or {}).get("nombre"), "created_at": __import__("datetime").datetime.now().isoformat(timespec="seconds"), "databases": ["auth_db", "documents_db", "catalog_db", "notifications_db"]},
    )
    headers = {"Content-Disposition": resp.headers.get("Content-Disposition", 'attachment; filename="respaldo.tar.gz"')}
    return Response(resp.iter_content(64 * 1024), content_type="application/gzip", headers=headers)


@app.post("/admin/respaldo/importar")
@admin_required
def importar_db():
    uploaded = request.files.get("backup_file") or request.files.get("file")
    if not uploaded:
        flash("Debe seleccionar un archivo de respaldo.", "warning")
        return redirect(url_for("admin_respaldo"))
    profile = current_user_profile()
    files = {"backup": (uploaded.filename, uploaded.stream, uploaded.mimetype)}
    resp = api("POST", BACKUP_URL, "/api/backups/import", files=files)
    if resp and resp.status_code == 200:
        write_system_log("BACKUP_RESTORE", "backup", detail={"filename": uploaded.filename})
        send_system_notification(
            "backup_restored",
            (profile or {}).get("email"),
            {"actor_name": (profile or {}).get("nombre"), "created_at": __import__("datetime").datetime.now().isoformat(timespec="seconds"), "filename": uploaded.filename},
        )
        session.clear()
        flash("Respaldo restaurado. Inicie sesión nuevamente.", "success")
        return redirect(url_for("index"))
    flash(response_error(resp, "No fue posible restaurar el respaldo"), "danger")
    return redirect(url_for("admin_respaldo"))


@app.get("/admin/logs")
@admin_required
def admin_logs():
    page = max(request.args.get("page", 1, type=int), 1)
    level = request.args.get("level", "").strip()
    module = request.args.get("module", "").strip()
    q = request.args.get("q", "").strip()
    resp = api("GET", NOTIFICATION_URL, "/api/system-logs", params={"page": page, "per_page": 50, "level": level, "module": module, "q": q})
    data = resp.json() if resp and resp.status_code == 200 else {"items": [], "summary": {}, "modules": [], "total": 0}
    return render_template("admin_logs.html", logs=data.get("items", []), summary=data.get("summary", {}), modules=data.get("modules", []), total=data.get("total", 0), page=page, level=level, module=module, q=q)


@app.get("/admin/logs/descargar")
@admin_required
def descargar_admin_logs():
    resp = api("GET", NOTIFICATION_URL, "/api/system-logs", params={"page": 1, "per_page": "all"})
    data = resp.json() if resp and resp.status_code == 200 else {"items": []}
    wb = Workbook()
    ws = wb.active
    ws.title = "Logs del sistema"
    headers = ["Fecha", "Nivel", "Módulo", "Evento", "Usuario", "Rol", "IP", "Detalle"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for item in data.get("items", []):
        import json as _json
        ws.append([item.get("created_at"), item.get("level"), item.get("module"), item.get("event_type"), item.get("actor_name"), item.get("actor_role"), item.get("ip_address"), _json.dumps(item.get("detail") or {}, ensure_ascii=False)])
    for idx, width in enumerate([24, 12, 18, 28, 28, 14, 18, 70], 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    write_system_log("SYSTEM_LOG_EXPORT", "reports", detail={"rows": len(data.get("items", []))})
    return send_file(output, as_attachment=True, download_name=f"logs_sistema_{datetime.now().date().isoformat()}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



@app.get("/health")
def health():
    checks = {}
    for name, url in {"auth": AUTH_URL, "documents": DOCUMENT_URL, "catalogs": CATALOG_URL, "backup": BACKUP_URL, "notifications": NOTIFICATION_URL}.items():
        resp = api("GET", url, "/health")
        checks[name] = bool(resp and resp.status_code == 200)
    return jsonify(status="ok" if all(checks.values()) else "degraded", services=checks), (200 if all(checks.values()) else 503)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000)
