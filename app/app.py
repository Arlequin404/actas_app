# app.py
from flask import Flask, render_template, request, redirect, session, url_for, flash
from markupsafe import Markup
import os
from dotenv import load_dotenv
from datetime import datetime, timedelta
import pytz
import psycopg2
from psycopg2.extras import RealDictCursor
import smtplib
from email.mime.text import MIMEText
from email.utils import formataddr
import secrets
from flask import send_file
from io import BytesIO
from datetime import datetime
import pandas as pd
import json
import subprocess
import tempfile

# ===============================
# Aplicación Flask
# ===============================

from sqlalchemy import create_engine

# ===============================
# Aplicación Flask & Config
# ===============================
load_dotenv()
app = Flask(__name__)
app.debug = False
app.secret_key = os.getenv("SECRET_KEY", "por_favor_cambia_esta_clave")

# ===============================
# Carga de entorno y configuración
# ===============================
APP_TZ = os.getenv("TZ", "America/Guayaquil")

# ===============================
# SMTP
# ===============================
SMTP_SERVER = os.getenv("SMTP_SERVER", "")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587") or 587)
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASS = os.getenv("SMTP_PASS", "")
SMTP_TLS = str(os.getenv("SMTP_TLS", "true")).lower() in {"1", "true", "yes", "on"}
SMTP_FROM = os.getenv("SMTP_FROM", "")  # opcional (si no se especifica, usa SMTP_USER)
APP_BASE_URL = os.getenv("APP_BASE_URL", "https://reportesdtcd.arconel.gob.ec")

# DATABASE_URL primero; fallback a POSTGRES_*
DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    pg_host = os.getenv("POSTGRES_HOST", os.getenv("POSTGRES_HOSTNAME", "db"))
    pg_port = int(os.getenv("POSTGRES_PORT", "5432") or 5432)
    pg_db = os.getenv("POSTGRES_DB", "actas_db")
    pg_user = os.getenv("POSTGRES_USER", "postgres")
    pg_pass = os.getenv("POSTGRES_PASSWORD", "postgres")
    DATABASE_URL = f"postgresql://{pg_user}:{pg_pass}@{pg_host}:{pg_port}/{pg_db}"

# SQLAlchemy engine for Pandas
engine = create_engine(DATABASE_URL)


# ===============================
# Utilidades
# ===============================
def get_conn():
    """Nueva conexión por request (seguro para Gunicorn)."""
    return psycopg2.connect(DATABASE_URL)

def password_matches(db_password: str, provided: str) -> bool:
    return db_password == provided


def send_email(to_email: str, subject: str, html: str):
    if not SMTP_SERVER:
        print("[WARN] SMTP no configurado; no se envía correo.")
        return
    msg = MIMEText(html, "html", "utf-8")
    sender = SMTP_FROM or SMTP_USER or "no-reply@example.com"
    if "<" in sender and ">" in sender:
        # ya viene con "Nombre <correo>"
        msg["From"] = sender
    else:
        msg["From"] = formataddr(("Notificaciones", sender))
    msg["To"] = to_email
    msg["Subject"] = subject

    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        if SMTP_TLS:
            server.starttls()
        if SMTP_USER:
            server.login(SMTP_USER, SMTP_PASS)
        server.send_message(msg)


def local_time_str(fecha, hora, tz_name=APP_TZ):
    """Convierte (fecha, hora) UTC a zona local y devuelve HH:MM:SS."""
    utc = pytz.utc
    local_tz = pytz.timezone(tz_name)
    dt_utc = datetime.combine(fecha, hora).replace(tzinfo=utc)
    dt_local = dt_utc.astimezone(local_tz)
    return dt_local.strftime("%H:%M:%S")

# ===============================
# Middlewares de sesión
# ===============================
def require_login():
    uid = session.get("user_id")
    if not uid:
        flash("Inicia sesión primero.", "warning")
        return False
    
    # Verificar si el usuario aún existe en la BD (por si hubo reset de BD)
    try:
        with get_conn() as conn, conn.cursor() as cur:
            cur.execute("SELECT 1 FROM usuarios WHERE id=%s", (uid,))
            if not cur.fetchone():
                session.clear()
                flash("Sesión inválida (usuario no encontrado). Inicia sesión de nuevo.", "warning")
                return False
    except Exception:
        # Si falla la BD, asumimos error pero no crasheamos aqui
        return False

    return True

def require_admin():
    if not require_login():
        return False
    if session.get("rol") != "admin":
        flash("Acceso denegado", "danger")
        return False
    return True

# ===============================
# Helpers Catalogos
# ===============================
def get_catalogo(categoria):
    """Retorna lista plana de opciones activas para una categoría dada."""
    with get_conn() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute("""
            SELECT id, nombre, valor, padre_id, meta_data
            FROM catalogos
            WHERE categoria = %s AND activo = TRUE
            ORDER BY orden ASC
        """, (categoria,))
        return cur.fetchall()

def get_all_companies():
    """Obtiene empresas del catálogo + empresas únicas encontradas en documentos (manualmente ingresadas)."""
    with get_conn() as conn, conn.cursor() as cur:
        # 1. Empresas del catálogo
        cur.execute("SELECT nombre, valor FROM catalogos WHERE categoria = 'EMPRESA' AND activo = TRUE ORDER BY orden")
        catalogo = cur.fetchall()
        
        empresas_dict = {}
        for c in catalogo:
            nombre = c[0]
            valor = c[1] if c[1] else c[0]
            empresas_dict[valor] = nombre
            
        # 2. Empresas en tablas de documentos
        tablas = ['actas', 'informes', 'reportes', 'comisiones']
        for t in tablas:
            cur.execute(f"SELECT DISTINCT empresa FROM {t} WHERE empresa IS NOT NULL AND empresa != ''")
            for row in cur.fetchall():
                emp = row[0]
                if emp not in empresas_dict:
                    empresas_dict[emp] = emp # Si no está en catálogo, usamos su nombre como valor
                    
    # Convertir a lista de dicts para compatibilidad con el template
    res = []
    # Ordenar alfabéticamente
    for valor in sorted(empresas_dict.keys()):
        res.append({
            'nombre': empresas_dict[valor],
            'valor': valor
        })
    return res

def build_hierarchy(items, parent_id=None):
    """Construye estructura recursiva para el JS de Reportes."""
    result = []
    # Filtrar items hijos del parent_id actual
    children = [x for x in items if x['padre_id'] == parent_id]
    
    for child in children:
        node = {
            'label': child['nombre'],
            'value': child['valor'] or child['nombre']
        }
        
        # Metadatos (banderas especiales)
        if child['meta_data']:
            # Extraer claves si existen
            # psycopg2 con RealDictCursor devuelve dict, y meta_data es jsonb (dict)
            md = child['meta_data']
            if md.get('special'):
                node['special'] = md['special']

        # Recursión
        grand_children = build_hierarchy(items, child['id'])
        if grand_children:
            node['children'] = grand_children
            
        result.append(node)
    return result

def get_jerarquia_informes():
    """Obtiene toda la categoría TIPO_INFORME y la estructura para el frontend."""
    with get_conn() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute("""
            SELECT id, nombre, valor, padre_id, meta_data 
            FROM catalogos 
            WHERE categoria = 'TIPO_INFORME' AND activo = TRUE 
            ORDER BY orden ASC
        """)
        all_items = cur.fetchall()
    
    # Construir árbol empezando por los que tienen padre_id IS NULL
    return build_hierarchy(all_items, None)


# ===============================
# Rutas
# ===============================
@app.route('/')
def index():
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def login():
    email = request.form.get('email', '').strip().lower()
    password = request.form.get('password', '')

    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT id, nombre, rol, password FROM usuarios WHERE LOWER(email)=%s", (email,))
        row = cur.fetchone()

    if row and password_matches(row[3], password):
        session['user_id'] = row[0]
        session['nombre'] = row[1]
        session['rol'] = row[2]
        return redirect('/dashboard')

    flash("Correo o contraseña incorrecta", "danger")
    return redirect('/')

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/')

@app.route('/dashboard')
def dashboard():
    if not require_login():
        return redirect('/')
    return render_template('dashboard.html', nombre=session['nombre'])

# ==============================================
# Exportar documentos (actas, informes, reportes, comisiones)
# ==============================================
@app.route('/exportar_documentos/<tipo>', methods=['GET'])
def exportar_documentos(tipo):
    if not require_login():
        return redirect('/')

    tablas_permitidas = ['actas', 'informes', 'reportes', 'comisiones']

    if tipo not in tablas_permitidas:
        flash("Tipo de documento no válido", "danger")
        return redirect('/dashboard')

    with get_conn() as conn:

        # ===============================
        # INFORMES (incluye CASOS FORTUITOS + gestiones + productos_asociados)
        # ===============================
        if tipo == 'informes':
            df = pd.read_sql_query("""
                SELECT
                    'INF.DTCD.' || i.anio || '.' || LPAD(i.numero::text, 3, '0') AS codigo,
                    i.empresa,
                    i.tipo_informe,

                    -- Gestiones / Productos asociados
                    i.gestiones,
                    i.productos_asociados,

                    -- Casos fortuitos
                    i.caso_tipo,
                    i.nombre_alimentador,
                    i.alimentador_subestacion,
                    i.linea_subtransmision_nombre,
                    i.fecha_interrupcion,

                    -- Datos generales
                    i.asunto,
                    i.observaciones,
                    i.fecha,
                    TO_CHAR(i.hora, 'HH24:MI') AS hora,

                    u.nombre AS funcionario
                FROM informes i
                JOIN usuarios u ON u.id = i.id_usuario
                ORDER BY i.anio DESC, i.numero DESC
            """, engine)

        # ===============================
        # REPORTES (incluye tipo_reporte + gestiones + productos_asociados)
        # ===============================
        elif tipo == 'reportes':
            df = pd.read_sql_query("""
                SELECT
                    'REP.DTCD.' || r.anio || '.' || LPAD(r.numero::text, 3, '0') AS codigo,
                    r.empresa,

                    -- NUEVO
                    r.gestiones,
                    r.productos_asociados,

                    r.tipo_reporte,
                    r.asunto,
                    r.observaciones,
                    r.fecha,
                    TO_CHAR(r.hora, 'HH24:MI') AS hora,
                    u.nombre AS funcionario
                FROM reportes r
                JOIN usuarios u ON u.id = r.id_usuario
                ORDER BY r.anio DESC, r.numero DESC
            """, engine)

        # ===============================
        # ACTAS / COMISIONES
        # ===============================
        else:
            df = pd.read_sql_query(f"""
                SELECT
                    CASE 
                        WHEN '{tipo}' = 'actas' THEN 'ACTAS.DTCD.' 
                        WHEN '{tipo}' = 'comisiones' THEN 'CMS.DTCD.' 
                        ELSE UPPER('{tipo}') || '.DTCD.' 
                    END || d.anio || '.' || LPAD(d.numero::text, 3, '0') AS codigo,
                    d.empresa,
                    d.gestiones,
                    d.productos_asociados,
                    d.asunto,
                    d.observaciones,
                    d.fecha,
                    TO_CHAR(d.hora, 'HH24:MI') AS hora,
                    u.nombre AS funcionario
                FROM {tipo} d
                JOIN usuarios u ON u.id = d.id_usuario
                ORDER BY d.anio DESC, d.numero DESC
            """, engine)

    # ===============================
    # Generar Excel con formato
    # ===============================
    
    # Renombrar columnas a español
    column_names_map = {
        'codigo': 'Código',
        'empresa': 'Empresa',
        'tipo_informe': 'Tipo de Informe',
        'tipo_reporte': 'Tipo de Reporte',
        'gestiones': 'Gestiones',
        'productos_asociados': 'Productos Asociados',
        'caso_tipo': 'Elemento Afectado',
        'nombre_alimentador': 'Nombre de Alimentador',
        'alimentador_subestacion': 'Subestación',
        'linea_subtransmision_nombre': 'Línea Subtransmisión',
        'fecha_interrupcion': 'Fecha Interrupción',
        'asunto': 'Asunto',
        'observaciones': 'Observaciones',
        'fecha': 'Fecha Registro',
        'hora': 'Hora Registro',
        'funcionario': 'Funcionario'
    }
    
    df.rename(columns=column_names_map, inplace=True)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=tipo.capitalize(), startrow=0, startcol=0)
        
        # Obtener el workbook y worksheet
        workbook = writer.book
        worksheet = writer.sheets[tipo.capitalize()]
        
        # Convertir a tabla de Excel
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Definir el rango de la tabla (desde A1 hasta la última columna y fila con datos)
        max_row = len(df) + 1  # +1 por el header
        max_col = len(df.columns)
        
        # Crear nombre de tabla único
        table_name = f"Tabla{tipo.capitalize()}"
        table_ref = f"A1:{chr(64 + max_col)}{max_row}"
        
        # Crear y configurar tabla
        tab = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        worksheet.add_table(tab)
        
        # Estilo para encabezados
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Aplicar estilo a encabezados
        for col_num in range(1, max_col + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Ajustar ancho de columnas automáticamente
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    cell_value = str(cell.value) if cell.value else ""
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            
            # Establecer ancho con un mínimo de 12 y un máximo de 50
            adjusted_width = min(max(max_length + 2, 12), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Aplicar bordes y alineación a todas las celdas
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        
        for row in worksheet.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    output.seek(0)
    filename = f"{tipo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ===============================
# Recuperación de contraseña
# ===============================
@app.route('/recuperar_contrasena', methods=['GET'])
@app.route('/recuperar_contraseña', methods=['GET'], endpoint='recuperar_contraseña')
def recuperar_contrasena():
    return render_template('recuperar_contraseña.html')

@app.route('/enviar_recuperacion', methods=['POST'])
def enviar_recuperacion():
    email = request.form.get('email', '').strip().lower()
    if not email:
        flash("Ingresa tu correo.", "warning")
        return redirect('/recuperar_contrasena')

    # ¿Existe el usuario?
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT 1 FROM usuarios WHERE LOWER(email)=%s", (email,))
        exists = cur.fetchone() is not None

    if not exists:
        flash("Correo no registrado.", "warning")
        return redirect('/recuperar_contrasena')

    token = secrets.token_urlsafe(32)
    expires_at = datetime.utcnow() + timedelta(hours=1)

    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("""
            INSERT INTO password_resets (token, email, expires_at)
            VALUES (%s, %s, %s)
            ON CONFLICT (token) DO UPDATE SET email=EXCLUDED.email, expires_at=EXCLUDED.expires_at
        """, (token, email, expires_at))
        conn.commit()

    # construir URL con base conocida (APP_BASE_URL) para que funcione detrás de proxy
    reset_path = url_for('restablecer_contrasena', token=token)
    reset_url = APP_BASE_URL.rstrip('/') + reset_path

    html = f"""
    <p>Hola,</p>
    <p>Has solicitado cambiar tu contraseña. Haz clic en el siguiente enlace:</p>
    <p><a href="{reset_url}">{reset_url}</a></p>
    <p>Este enlace expira en 1 hora.</p>
    """
    try:
        send_email(email, "Recuperación de contraseña", html)
        flash("Se ha enviado un enlace de recuperación al correo.", "success")
    except Exception as e:
        print("Error al enviar correo:", e)
        flash("Error al enviar el correo.", "danger")

    return redirect('/')

@app.route('/restablecer_contrasena/<token>', methods=['GET', 'POST'])
@app.route('/restablecer_contraseña/<token>', methods=['GET', 'POST'], endpoint='restablecer_contraseña')
def restablecer_contrasena(token):
    if request.method == 'POST':
        nueva = request.form.get('password', '')
        email = request.form.get('email', '').strip().lower()

        if not nueva or not email:
            flash("Completa los campos requeridos.", "warning")
            return redirect(request.url)

        # validar token
        with get_conn() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT email, expires_at
                FROM password_resets
                WHERE token=%s
            """, (token,))
            row = cur.fetchone()

            if not row or row['email'].lower() != email or row['expires_at'] < datetime.utcnow():
                flash("Enlace inválido o expirado.", "danger")
                return redirect('/')

            # actualizar contraseña
            cur.execute("UPDATE usuarios SET password=%s WHERE LOWER(email)=%s", (nueva, email))

            # borrar token
            cur.execute("DELETE FROM password_resets WHERE token=%s", (token,))
            conn.commit()

        flash("Contraseña actualizada exitosamente.", "success")
        return redirect('/')

    # GET
    return render_template('restablecer_contraseña.html', token=token)

# ===============================
# ADMIN CONFIGURACION
# ===============================
@app.route('/admin/configuracion')
def admin_configuracion():
    if not require_admin():
        return redirect('/')
    return render_template('admin_configuracion.html')

# ===============================
# RESPALDO BASE DE DATOS
# ===============================

@app.route('/admin/respaldo')
def admin_respaldo():
    if not require_admin():
        return redirect('/')
    return render_template('admin_respaldo.html')

@app.route('/admin/respaldo/exportar')
def exportar_db():
    if not require_admin():
        return redirect('/')
    
    # Credenciales desde variables de entorno
    pg_host = os.getenv("POSTGRES_HOST", "db")
    pg_port = os.getenv("POSTGRES_PORT", "5432")
    pg_db = os.getenv("POSTGRES_DB", "actas_db")
    pg_user = os.getenv("POSTGRES_USER", "postgres")
    pg_pass = os.getenv("POSTGRES_PASSWORD", "postgres")

    # Usar PGPASSWORD env var para pg_dump
    env = os.environ.copy()
    env["PGPASSWORD"] = pg_pass

    filename = f"respaldo_{pg_db}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sql"
    
    try:
        # Comando pg_dump con opciones de limpieza para que el restore sea más fácil
        # --clean: DROP database objects before creating them
        # --if-exists: Use IF EXISTS when dropping objects
        # --no-owner: Do not output commands to set ownership of objects
        # --no-privileges: Do not output commands to set access privileges (grant/revoke)
        process = subprocess.run(
            [
                "pg_dump", 
                "-h", pg_host, 
                "-p", pg_port, 
                "-U", pg_user, 
                "-d", pg_db, 
                "--clean",
                "--if-exists",
                "--no-owner", 
                "--no-privileges"
            ],
            env=env,
            capture_output=True,
            text=True,
            check=True
        )
        backup_sql = process.stdout
        
        return send_file(
            BytesIO(backup_sql.encode('utf-8')),
            download_name=filename,
            as_attachment=True,
            mimetype='application/sql'
        )
    except subprocess.CalledProcessError as e:
        flash(f"Error al ejecutar backup: {e.stderr}", "danger")
        return redirect(url_for('admin_respaldo'))
    except Exception as e:
        flash(f"Error inesperado: {str(e)}", "danger")
        return redirect(url_for('admin_respaldo'))

@app.route('/admin/respaldo/importar', methods=['POST'])
def importar_db():
    if not require_admin():
        return redirect('/')
    
    file = request.files.get('backup_file')
    if not file or file.filename == '':
        flash("No se seleccionó ningún archivo", "warning")
        return redirect(url_for('admin_respaldo'))

    # Credenciales desde variables de entorno
    pg_host = os.getenv("POSTGRES_HOST", "db")
    pg_port = os.getenv("POSTGRES_PORT", "5432")
    pg_db = os.getenv("POSTGRES_DB", "actas_db")
    pg_user = os.getenv("POSTGRES_USER", "postgres")
    pg_pass = os.getenv("POSTGRES_PASSWORD", "postgres")

    env = os.environ.copy()
    env["PGPASSWORD"] = pg_pass

    try:
        # Guardar en archivo temporal
        with tempfile.NamedTemporaryFile(suffix=".sql", delete=False) as tf:
            file.save(tf)
            tf_path = tf.name

        # Para asegurar una restauración limpia, primero vaciamos el esquema public.
        # Esto soluciona los errores de "relation already exists".
        # ADVERTENCIA: Esto borra TODO en el esquema public antes de restaurar.
        try:
            with get_conn() as conn, conn.cursor() as cur:
                cur.execute("DROP SCHEMA public CASCADE; CREATE SCHEMA public;")
                conn.commit()
                print("[INFO] Esquema public limpiado para restauración.")
        except Exception as e:
            print(f"[ERROR] No se pudo limpiar el esquema public: {e}")
            # Continuamos de todos modos, psql intentará lo mejor posible (especialmente con dumps --clean)

        # Ejecutar psql
        process = subprocess.run(
            ["psql", "-h", pg_host, "-p", pg_port, "-U", pg_user, "-d", pg_db, "-f", tf_path],
            env=env,
            capture_output=True,
            text=True
        )
        
        # El psql puede fallar en algunos comandos (como SET transaction_timeout)
        # pero aun así restaurar la mayoría de los datos.
        
        os.remove(tf_path)
        
        if process.returncode == 0:
            flash("Copia seguridad importada correctamente.", "success")
        else:
            # Si hay errores, los mostramos pero a menudo la base se restaura parcialmente
            error_msg = process.stderr
            print(f"[DEBUG] psql output: {process.stdout}")
            print(f"[DEBUG] psql errors: {error_msg}")
            
            if "already exists" in error_msg or "violates unique constraint" in error_msg:
                flash("La restauración terminó con algunas advertencias (objetos duplicados). Verifica los datos.", "warning")
            else:
                flash(f"Error al importar archivo SQL: {error_msg[:200]}...", "danger")
            
    except Exception as e:
        flash(f"Error en el proceso de importación: {str(e)}", "danger")

    return redirect(url_for('admin_respaldo'))

@app.route('/api/catalogos/<categoria>', methods=['GET'])
def api_get_catalogos(categoria):
    if not require_login(): # ¿Permitir lectura a usuarios logueados o solo admin?
        return {"error": "No autorizado"}, 401
    
    # Si piden jerarquía completa
    if categoria == 'TIPO_INFORME_TREE':
        return {"data": get_jerarquia_informes()}

    with get_conn() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute("""
            SELECT id, categoria, nombre, valor, padre_id, activo, orden, meta_data 
            FROM catalogos 
            WHERE categoria = %s 
            ORDER BY orden ASC
        """, (categoria,))
        data = cur.fetchall()
    return {"data": data}

@app.route('/api/catalogos', methods=['POST'])
def api_create_catalogo():
    if not require_admin():
        return {"error": "Unauthorized"}, 403
    
    data = request.json
    try:
        with get_conn() as conn, conn.cursor() as cur:
            # Clean data
            padre_id = data.get('padre_id')
            if padre_id == "":
                padre_id = None
            
            orden = data.get('orden', 0)
            try:
                orden = int(orden)
            except:
                orden = 0

            cur.execute("""
                INSERT INTO catalogos (categoria, nombre, valor, padre_id, orden, meta_data)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                data.get('categoria'),
                data.get('nombre'),
                data.get('valor'),
                padre_id, 
                orden,
                json.dumps(data.get('meta_data')) if data.get('meta_data') else None
            ))
            new_id = cur.fetchone()[0]
            conn.commit()
        return {"success": True, "id": new_id}
    except Exception as e:
        print(f"Error al crear catálogo: {e}")
        return {"error": str(e)}, 500

@app.route('/api/catalogos/<int:id>', methods=['PUT'])
def api_update_catalogo(id):
    if not require_admin():
        return {"error": "Unauthorized"}, 403
    
    data = request.json
    try:
        with get_conn() as conn, conn.cursor() as cur:
            # Clean data
            padre_id = data.get('padre_id')
            if padre_id == "":
                padre_id = None
            
            orden = data.get('orden', 0)
            try:
                orden = int(orden)
            except:
                orden = 0

            cur.execute("""
                UPDATE catalogos 
                SET nombre=%s, valor=%s, padre_id=%s, orden=%s, activo=%s, meta_data=%s
                WHERE id=%s
            """, (
                data.get('nombre'),
                data.get('valor'),
                padre_id,
                orden,
                data.get('activo', True),
                json.dumps(data.get('meta_data')) if data.get('meta_data') else None,
                id
            ))
            conn.commit()
        return {"success": True}
    except Exception as e:
        print(f"Error al actualizar catálogo: {e}")
        return {"error": str(e)}, 500

@app.route('/api/catalogos/<int:id>', methods=['DELETE'])
def api_delete_catalogo(id):
    if not require_admin():
        return {"error": "Unauthorized"}, 403
    
    try:
        with get_conn() as conn, conn.cursor() as cur:
            cur.execute("DELETE FROM catalogos WHERE id=%s", (id,))
            conn.commit()
        return {"success": True}
    except Exception as e:
        return {"error": str(e)}, 500

# ===============================
# Creación de documentos
# ===============================

@app.route('/crear/<tipo>', methods=['GET', 'POST'])
def crear(tipo):
    if not require_login():
        return redirect('/')

    if session.get('rol') not in ['usuario', 'admin']:
        return 'Acceso no autorizado'

    tipo = tipo.lower()

    tablas_permitidas = {
        'actas': 'actas',
        'informes': 'informes',
        'reportes': 'reportes',
        'comisiones': 'comisiones'
    }

    if tipo not in tablas_permitidas:
        flash("Tipo de documento inválido.", "danger")
        return redirect('/dashboard')

    tabla = tablas_permitidas[tipo]

    # Datos para los selectores
    empresas = get_catalogo('EMPRESA')
    
    # Listas vacías por defecto
    gestiones = []
    productos = []
    report_types_json = []
    productos = []
    report_types_json = []
    tipos_reporte = []
    
    if tipo == 'informes':
        gestiones = get_catalogo('GESTION_INFORME')
        productos = get_catalogo('PRODUCTO_INFORME')
        report_types_json = get_jerarquia_informes()


        
    elif tipo == 'reportes':
        gestiones = get_catalogo('GESTION_REPORTE')
        productos = get_catalogo('PRODUCTO_REPORTE')
        tipos_reporte = get_catalogo('TIPO_REPORTE')

    elif tipo == 'actas':
        gestiones = get_catalogo('GESTION_ACTA')
        productos = get_catalogo('PRODUCTO_ACTA')

    elif tipo == 'comisiones':
        gestiones = get_catalogo('GESTION_COMISION')
        productos = get_catalogo('PRODUCTO_COMISION')

    if request.method == 'POST':
        numero_manual = request.form.get('numero_manual', '').strip()
        try:
            numero_manual = int(numero_manual) if numero_manual else None
        except:
            numero_manual = None

        asunto = request.form.get('asunto', '').strip()
        observaciones = request.form.get('observaciones', '').strip()
        empresa = request.form.get('empresa', '').strip()
        empresa_otro = request.form.get('empresa_otro', '').strip()

        if empresa == "Otros":
            if not empresa_otro:
                flash("Debe especificar la empresa cuando selecciona 'Otros'.", "warning")
                return redirect(request.url)
            empresa = empresa_otro

        # ===============================
        # INFORMES / ACTAS / COMISIONES
        # ===============================
        # Ahora todos pueden tener Gestiones / Productos

        tipo_informe = request.form.get('tipo_informe', '').strip()
        tipo_informe_otro = request.form.get('tipo_informe_otro', '').strip()

        # GESTIONES y PRODUCTOS (para informes, actas, comisiones)
        # Reutilizamos las mismas variables del form si los 'name' coinciden
        # En el HTML usaremos 'gestiones' y 'productos_asociados' para estos 3 tipos.
        gestiones = request.form.get('gestiones', '').strip()
        gestiones_otro = request.form.get('gestiones_otro', '').strip()

        productos_asociados = request.form.get('productos_asociados', '').strip()
        productos_asociados_otro = request.form.get('productos_asociados_otro', '').strip()

        # SOLO REPORTES
        tipo_reporte = request.form.get('tipo_reporte', '').strip()
        tipo_reporte_otro = request.form.get('tipo_reporte_otro', '').strip()

        # REPORTES: GESTIONES / PRODUCTOS ASOCIADOS (NUEVO)
        gestiones_reporte = request.form.get('gestiones_reporte', '').strip()
        gestiones_reporte_otro = request.form.get('gestiones_reporte_otro', '').strip()

        productos_asociados_reporte = request.form.get('productos_asociados_reporte', '').strip()
        productos_asociados_reporte_otro = request.form.get('productos_asociados_reporte_otro', '').strip()

        # ===============================
        # CAMPOS (CASO FORTUITO)
        # ===============================
        caso_tipo = request.form.get('caso_tipo', '').strip()
        nombre_alimentador = request.form.get('nombre_alimentador', '').strip()
        alimentador_subestacion = request.form.get('alimentador_subestacion', '').strip()
        linea_subtransmision_nombre = request.form.get('linea_subtransmision_nombre', '').strip()

        fecha_interrupcion = request.form.get('fecha_interrupcion', '').strip()  # 'YYYY-MM-DD' o ''

        if not asunto:
            flash("El asunto es obligatorio.", "warning")
            return redirect(request.url)

        if not empresa:
            flash("Debe seleccionar una empresa.", "warning")
            return redirect(request.url)

        # Convertir a None si está vacío (para que Postgres guarde NULL)
        def empty_to_none(x):
            x = (x or "").strip()
            return x if x else None

        # ===============================
        # Validación + reemplazo cuando es "Otros"
        # ===============================
        # ===============================
        # Resolver "Otros" para todos los tipos
        # ===============================
        if gestiones == "Otros":
            if not gestiones_otro:
                flash("Debe especificar la gestión cuando selecciona 'Otros'.", "warning")
                return redirect(request.url)
            gestiones = gestiones_otro

        if productos_asociados == "Otros":
            if not productos_asociados_otro:
                flash("Debe especificar el producto asociado cuando selecciona 'Otros'.", "warning")
                return redirect(request.url)
            productos_asociados = productos_asociados_otro

        if tipo_informe == "Otros":
            if not tipo_informe_otro:
                flash("Debe especificar el tipo de informe cuando selecciona 'Otros'.", "warning")
                return redirect(request.url)
            tipo_informe = tipo_informe_otro

        if tipo == 'informes':
            if not tipo_informe:
                flash("Debe seleccionar el tipo de informe.", "warning")
                return redirect(request.url)
            if not gestiones:
                flash("Debe seleccionar una gestión.", "warning")
                return redirect(request.url)
            if not productos_asociados:
                flash("Debe seleccionar un producto asociado.", "warning")
                return redirect(request.url)
        
        elif tipo in ['actas', 'comisiones']:
            if not gestiones:
                flash("Debe seleccionar una gestión.", "warning")
                return redirect(request.url)
            if not productos_asociados:
                flash("Debe seleccionar un producto asociado.", "warning")
                return redirect(request.url)

            # ===============================
            # VALIDACIÓN CASO FORTUITO
            # ===============================
            # Verificar si existe la subcadena porque el nuevo valor es más largo (insensible a mayúsculas)
            if "caso fortuito" in tipo_informe.lower():
                if not caso_tipo:
                    flash("En 'Caso Fortuito' debe seleccionar: ALIMENTADOR o LINEAS DE SUBTRANSMISION.", "warning")
                    return redirect(request.url)

                if not fecha_interrupcion:
                    flash("En 'Caso Fortuito' debe seleccionar la FECHA de interrupción.", "warning")
                    return redirect(request.url)

                if caso_tipo == "ALIMENTADOR":
                    if not nombre_alimentador:
                        flash("En 'ALIMENTADOR' debe ingresar el nombre del alimentador.", "warning")
                        return redirect(request.url)
                    if not alimentador_subestacion:
                        flash("En 'ALIMENTADOR' debe ingresar la subestación.", "warning")
                        return redirect(request.url)

                    # limpiar campos del otro tipo
                    linea_subtransmision_nombre = ""

                elif caso_tipo == "LINEAS DE SUBTRANSMISION":
                    if not linea_subtransmision_nombre:
                        flash("En 'LINEAS DE SUBTRANSMISION' debe ingresar el nombre.", "warning")
                        return redirect(request.url)

                    # limpiar campos del otro tipo
                    nombre_alimentador = ""
                    alimentador_subestacion = ""

                else:
                    flash("Selección inválida en Caso Tipo.", "warning")
                    return redirect(request.url)

            else:
                # si NO es Caso Fortuito, guardar como NULL
                caso_tipo = ""
                nombre_alimentador = ""
                alimentador_subestacion = ""
                linea_subtransmision_nombre = ""
                fecha_interrupcion = ""

        if tipo == 'reportes':
            if not tipo_reporte:
                flash("Debe seleccionar el tipo de reporte.", "warning")
                return redirect(request.url)

            if tipo_reporte == "Otros":
                if not tipo_reporte_otro:
                    flash("Debe especificar el tipo de reporte cuando selecciona 'Otros'.", "warning")
                    return redirect(request.url)
                tipo_reporte = tipo_reporte_otro

            # NUEVO: Gestiones / Productos asociados obligatorios en reportes
            if not gestiones_reporte:
                flash("Debe seleccionar una gestión (Reportes).", "warning")
                return redirect(request.url)

            if not productos_asociados_reporte:
                flash("Debe seleccionar un producto asociado (Reportes).", "warning")
                return redirect(request.url)

            if gestiones_reporte == "Otros":
                if not gestiones_reporte_otro:
                    flash("Debe especificar la gestión cuando selecciona 'Otros' (Reportes).", "warning")
                    return redirect(request.url)
                gestiones_reporte = gestiones_reporte_otro

            if productos_asociados_reporte == "Otros":
                if not productos_asociados_reporte_otro:
                    flash("Debe especificar el producto asociado cuando selecciona 'Otros' (Reportes).", "warning")
                    return redirect(request.url)
                productos_asociados_reporte = productos_asociados_reporte_otro

        timezone = pytz.timezone(APP_TZ)

        now = datetime.now(timezone)
        fecha = now.date()
        hora = now.time()
        anio = fecha.year

        # ===============================
        # Preparar valores DB
        # ===============================
        caso_tipo_db = empty_to_none(caso_tipo)
        nombre_alimentador_db = empty_to_none(nombre_alimentador)
        alimentador_subestacion_db = empty_to_none(alimentador_subestacion)
        linea_subtransmision_nombre_db = empty_to_none(linea_subtransmision_nombre)

        fecha_interrupcion_db = empty_to_none(fecha_interrupcion)

        gestiones_db = empty_to_none(gestiones)
        productos_asociados_db = empty_to_none(productos_asociados)

        gestiones_reporte_db = empty_to_none(gestiones_reporte)
        productos_asociados_reporte_db = empty_to_none(productos_asociados_reporte)

        with get_conn() as conn, conn.cursor() as cur:

            # ===============================
            # INSERT SEGÚN TIPO
            # ===============================
            if tipo == 'informes':
                cur.execute("""
                    INSERT INTO informes (
                        numero,
                        anio,
                        empresa,
                        tipo_informe,

                        gestiones,
                        productos_asociados,

                        asunto,
                        observaciones,
                        id_usuario,
                        fecha,
                        hora,

                        caso_tipo,
                        nombre_alimentador,
                        alimentador_subestacion,
                        linea_subtransmision_nombre,
                        fecha_interrupcion
                    )
                    VALUES (
                        COALESCE(%s, (
                            SELECT COALESCE(MAX(numero), 0) + 1
                            FROM informes
                            WHERE anio = %s
                        )),
                        %s, %s, %s,
                        %s, %s,
                        %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s
                    )
                    RETURNING id, numero, anio
                """, (
                    numero_manual,
                    anio,
                    anio,
                    empresa,
                    tipo_informe,

                    gestiones_db,
                    productos_asociados_db,

                    asunto,
                    observaciones,
                    session['user_id'],
                    fecha,
                    hora,

                    caso_tipo_db,
                    nombre_alimentador_db,
                    alimentador_subestacion_db,
                    linea_subtransmision_nombre_db,
                    fecha_interrupcion_db
                ))

            elif tipo == 'reportes':
                cur.execute("""
                    INSERT INTO reportes (
                        numero,
                        anio,
                        empresa,

                        gestiones,
                        productos_asociados,

                        tipo_reporte,
                        asunto,
                        observaciones,
                        id_usuario,
                        fecha,
                        hora
                    )
                    VALUES (
                        COALESCE(%s, (
                            SELECT COALESCE(MAX(numero), 0) + 1
                            FROM reportes
                            WHERE anio = %s
                        )),
                        %s, %s,
                        %s, %s,
                        %s, %s, %s, %s, %s, %s
                    )
                    RETURNING id, numero, anio
                """, (
                    numero_manual,
                    anio,
                    anio,
                    empresa,

                    gestiones_reporte_db,
                    productos_asociados_reporte_db,

                    tipo_reporte,
                    asunto,
                    observaciones,
                    session['user_id'],
                    fecha,
                    hora
                ))

            else:
                # Actas y Comisiones ahora tienen gestiones y productos
                cur.execute(f"""
                    INSERT INTO {tabla} (
                        numero,
                        anio,
                        empresa,
                        asunto,
                        observaciones,
                        id_usuario,
                        fecha,
                        hora,
                        gestiones, 
                        productos_asociados
                    )
                    VALUES (
                        COALESCE(%s, (
                            SELECT COALESCE(MAX(numero), 0) + 1
                            FROM {tabla}
                            WHERE anio = %s
                        )),
                        %s, %s, %s, %s, %s, %s, %s, %s, %s
                    )
                    RETURNING id, numero, anio
                """, (
                    numero_manual,
                    anio,
                    anio,
                    empresa,
                    asunto,
                    observaciones,
                    session['user_id'],
                    fecha,
                    hora,
                    gestiones_db,         # Agregado
                    productos_asociados_db # Agregado
                ))

            doc_id, numero, anio = cur.fetchone()

            cur.execute(
                "SELECT email FROM usuarios WHERE id=%s",
                (session['user_id'],)
            )
            email_usuario = cur.fetchone()[0]

            conn.commit()

        hora_local = hora.strftime("%H:%M:%S")

        # ===============================
        # Prefijos / Código
        # ===============================
        prefijos = {
            "informes": "INF.DTCD",
            "actas": "ACTAS.DTCD",
            "reportes": "REP.DTCD",
            "comisiones": "CMS.DTCD"
        }
        prefijo = prefijos.get(tipo, tipo.upper())
        codigo = f"{prefijo}.{anio}.{str(numero).zfill(3)}"

        # ===============================
        # Correo
        # ===============================
        try:
            # Construir HTML mejorado con tabla
            html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    body {{ font-family: Arial, sans-serif; color: #333; }}
                    .container {{ max-width: 800px; margin: 0 auto; padding: 20px; }}
                    .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; border-radius: 8px 8px 0 0; }}
                    .header h2 {{ margin: 0; font-size: 24px; }}
                    .content {{ background: #f9f9f9; padding: 20px; border-radius: 0 0 8px 8px; }}
                    .info-table {{ width: 100%; border-collapse: collapse; margin-top: 15px; background: white; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
                    .info-table th {{ background: #4a5568; color: white; padding: 12px; text-align: left; font-weight: 600; }}
                    .info-table td {{ padding: 12px; border-bottom: 1px solid #e2e8f0; }}
                    .info-table tr:last-child td {{ border-bottom: none; }}
                    .info-table tr:nth-child(even) {{ background: #f7fafc; }}
                    .label {{ font-weight: 600; color: #4a5568; width: 200px; }}
                    .section-title {{ background: #edf2f7; padding: 10px; margin-top: 20px; font-weight: 700; color: #2d3748; border-left: 4px solid #667eea; }}
                    .footer {{ margin-top: 20px; padding: 15px; background: #edf2f7; border-radius: 8px; font-size: 12px; color: #718096; text-align: center; }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h2>✓ Nuevo {tipo.capitalize()} Registrado</h2>
                    </div>
                    <div class="content">
                        <table class="info-table">
                            <tr>
                                <td class="label">Código</td>
                                <td><strong style="color: #667eea; font-size: 16px;">{codigo}</strong></td>
                            </tr>
                            <tr>
                                <td class="label">Empresa</td>
                                <td>{empresa}</td>
                            </tr>
            """
            
            # Campos específicos según tipo de documento
            if tipo == "informes":
                html += f"""
                            <tr>
                                <td class="label">Tipo de Informe</td>
                                <td>{tipo_informe}</td>
                            </tr>
                            <tr>
                                <td class="label">Gestiones</td>
                                <td>{gestiones_db or 'N/A'}</td>
                            </tr>
                            <tr>
                                <td class="label">Productos Asociados</td>
                                <td>{productos_asociados_db or 'N/A'}</td>
                            </tr>
                """
            
            elif tipo == "reportes":
                html += f"""
                            <tr>
                                <td class="label">Tipo de Reporte</td>
                                <td>{tipo_reporte}</td>
                            </tr>
                            <tr>
                                <td class="label">Gestiones</td>
                                <td>{gestiones_reporte_db or 'N/A'}</td>
                            </tr>
                            <tr>
                                <td class="label">Productos Asociados</td>
                                <td>{productos_asociados_reporte_db or 'N/A'}</td>
                            </tr>
                """
            
            elif tipo in ["actas", "comisiones"]:
                html += f"""
                            <tr>
                                <td class="label">Gestiones</td>
                                <td>{gestiones_db or 'N/A'}</td>
                            </tr>
                            <tr>
                                <td class="label">Productos Asociados</td>
                                <td>{productos_asociados_db or 'N/A'}</td>
                            </tr>
                """
            
            # Campos comunes
            html += f"""
                            <tr>
                                <td class="label">Asunto</td>
                                <td>{asunto}</td>
                            </tr>
                            <tr>
                                <td class="label">Observaciones</td>
                                <td>{observaciones if observaciones else 'N/A'}</td>
                            </tr>
                            <tr>
                                <td class="label">Fecha de Registro</td>
                                <td>{fecha}</td>
                            </tr>
                            <tr>
                                <td class="label">Hora de Registro</td>
                                <td>{hora_local}</td>
                            </tr>
                            <tr>
                                <td class="label">Funcionario</td>
                                <td>{session['nombre']}</td>
                            </tr>
                        </table>
            """
            
            # Caso Fortuito (solo para informes)
            if tipo == "informes" and "caso fortuito" in tipo_informe.lower():
                html += f"""
                        <div class="section-title">📋 Detalles de Caso Fortuito</div>
                        <table class="info-table">
                            <tr>
                                <td class="label">Tipo</td>
                                <td>{caso_tipo_db or 'N/A'}</td>
                            </tr>
                            <tr>
                                <td class="label">Fecha de Interrupción</td>
                                <td>{fecha_interrupcion_db or 'N/A'}</td>
                            </tr>
                """
                
                if caso_tipo_db == "ALIMENTADOR":
                    html += f"""
                            <tr>
                                <td class="label">Alimentador</td>
                                <td>{nombre_alimentador_db or 'N/A'}</td>
                            </tr>
                            <tr>
                                <td class="label">Subestación</td>
                                <td>{alimentador_subestacion_db or 'N/A'}</td>
                            </tr>
                    """
                elif caso_tipo_db == "LINEAS DE SUBTRANSMISION":
                    html += f"""
                            <tr>
                                <td class="label">Línea de Subtransmisión</td>
                                <td>{linea_subtransmision_nombre_db or 'N/A'}</td>
                            </tr>
                    """
                
                html += """
                        </table>
                """
            
            html += """
                        <div class="footer">
                            Este es un correo automático generado por el Sistema de Gestión de Documentos DTCD.<br>
                            Por favor no responder a este correo.
                        </div>
                    </div>
                </div>
            </body>
            </html>
            """

            send_email(
                email_usuario,
                f"{codigo} creado",
                html
            )
        except Exception as e:
            print("Error al enviar correo:", e)

        # ===============================
        # Toast
        # ===============================
        flash(Markup(f"""
        <div class='toast-content'>
            <h5>{tipo.capitalize()} creado correctamente</h5>
            <p><b>Código:</b> {codigo}</p>
            <p><b>Empresa:</b> {empresa}</p>

            {"<p><b>Tipo de informe:</b> " + tipo_informe + "</p>" if tipo == "informes" else ""}
            {"<p><b>Gestiones:</b> " + (gestiones_db or '') + "</p>" if tipo == "informes" else ""}
            {"<p><b>Productos asociados:</b> " + (productos_asociados_db or '') + "</p>" if tipo == "informes" else ""}

            {"<p><b>Gestiones:</b> " + (gestiones_reporte_db or '') + "</p>" if tipo == "reportes" else ""}
            {"<p><b>Productos asociados:</b> " + (productos_asociados_reporte_db or '') + "</p>" if tipo == "reportes" else ""}
            {"<p><b>Tipo de reporte:</b> " + tipo_reporte + "</p>" if tipo == "reportes" else ""}

            {"<p><b>Gestiones:</b> " + (gestiones_db or '') + "</p>" if tipo in ["actas", "comisiones"] else ""}
            {"<p><b>Productos asociados:</b> " + (productos_asociados_db or '') + "</p>" if tipo in ["actas", "comisiones"] else ""}

            <p><b>Asunto:</b> {asunto}</p>
            {"<p><b>Observaciones:</b> " + observaciones + "</p>" if observaciones else ""}
            <p><b>Fecha:</b> {fecha}</p>
            <p><b>Hora:</b> {hora_local}</p>
        </div>
        """), 'toast')

        return redirect('/dashboard')

    return render_template(
        'formulario.html', 
        tipo=tipo, 
        nombre=session['nombre'],
        
        # Opciones dinámicas
        empresas=empresas,
        gestiones=gestiones,
        productos_asociados=productos,
        report_types_json=report_types_json,
        tipos_reporte=tipos_reporte
    )


# ===============================
# Panel de administración
# ===============================
@app.route('/admin')
def admin():
    if not require_admin():
        return redirect('/')

    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("""
            SELECT id, nombre, email, rol
            FROM usuarios
            ORDER BY id
        """)
        usuarios = cur.fetchall()

    return render_template('admin.html', usuarios=usuarios)


def format_doc_code(tipo, numero, anio):
    prefijos = {
        "informes": "INF.DTCD",
        "actas": "ACTAS.DTCD",
        "reportes": "REP.DTCD",
        "comisiones": "CMS.DTCD"
    }
    prefijo = prefijos.get(tipo, tipo.upper())
    return f"{prefijo}.{anio}.{str(numero).zfill(3)}"

def get_paginated_docs(tabla, empresa=None, user_id=None, page=1, per_page=10):
    """Auxiliar para obtener documentos paginados y filtrados."""
    
    # Handle 'all' per_page
    if per_page == 'all':
        limit = 999999
        offset = 0
        page = 1
    else:
        try:
            per_page = int(per_page)
        except:
            per_page = 10
        limit = per_page
        offset = (page - 1) * per_page

    where_parts = []
    params = []

    if empresa:
        where_parts.append("empresa = %s")
        params.append(empresa)
    
    if user_id:
        where_parts.append("id_usuario = %s")
        params.append(user_id)

    where_sql = ""
    if where_parts:
        where_sql = "WHERE " + " AND ".join(where_parts)

    # Columnas según tabla
    if tabla == 'informes':
        cols = """
            t.id, t.numero, t.anio, t.empresa, t.gestiones, t.productos_asociados, t.tipo_informe,
            t.caso_tipo, t.nombre_alimentador, t.alimentador_subestacion, t.linea_subtransmision_nombre,
            t.fecha_interrupcion, t.asunto, t.observaciones, t.fecha, TO_CHAR(t.hora, 'HH24:MI:SS') as hora
        """
    elif tabla == 'reportes':
        cols = """
            t.id, t.numero, t.anio, t.empresa, t.gestiones, t.productos_asociados, t.tipo_reporte,
            t.asunto, t.observaciones, t.fecha, TO_CHAR(t.hora, 'HH24:MI:SS') as hora
        """
    else: # actas, comisiones
        cols = """
            t.id, t.numero, t.anio, t.empresa, t.gestiones, t.productos_asociados,
            t.asunto, t.observaciones, t.fecha, TO_CHAR(t.hora, 'HH24:MI:SS') as hora
        """

    with get_conn() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
        # 1. Total count
        cur.execute(f"SELECT COUNT(*) FROM {tabla} {where_sql}", params)
        total = cur.fetchone()['count']

        # 2. Data
        cur.execute(f"""
            SELECT {cols}, u.nombre as funcionario
            FROM {tabla} t
            JOIN usuarios u ON u.id = t.id_usuario
            {where_sql}
            ORDER BY t.anio DESC, t.numero DESC
            LIMIT %s OFFSET %s
        """, params + [limit, offset])
        rows = cur.fetchall()

    if per_page == 'all':
        pages = 1
    else:
        pages = (total + per_page - 1) // per_page
    return rows, total, pages

@app.route('/admin/documentos')
def admin_documentos():
    if not require_admin():
        return redirect('/')

    per_page = request.args.get('per_page', '10')
    page = request.args.get('page', 1, type=int)
    if page < 1: page = 1
    empresa_filtro = request.args.get('empresa', '').strip()
    active_tab = request.args.get('tab', 'actas')

    actas_raw, total_a, pages_a = get_paginated_docs('actas', empresa_filtro, page=page, per_page=per_page)
    informes_raw, total_i, pages_i = get_paginated_docs('informes', empresa_filtro, page=page, per_page=per_page)
    reportes_raw, total_r, pages_r = get_paginated_docs('reportes', empresa_filtro, page=page, per_page=per_page)
    comisiones_raw, total_c, pages_c = get_paginated_docs('comisiones', empresa_filtro, page=page, per_page=per_page)

    def fmt_docs(docs, tipo):
        res = []
        for d in docs:
            codigo = format_doc_code(tipo, d['numero'], d['anio'])
            if tipo == 'informes':
                res.append((d['id'], codigo, d['empresa'], d['gestiones'], d['productos_asociados'], d['tipo_informe'], d['caso_tipo'], d['nombre_alimentador'], d['alimentador_subestacion'], d['linea_subtransmision_nombre'], d['fecha_interrupcion'], d['asunto'], d['observaciones'], d['fecha'], d['hora'], d['funcionario']))
            elif tipo == 'reportes':
                res.append((d['id'], codigo, d['empresa'], d['gestiones'], d['productos_asociados'], d['tipo_reporte'], d['asunto'], d['observaciones'], d['fecha'], d['hora'], d['funcionario']))
            else:
                res.append((d['id'], codigo, d['empresa'], d['gestiones'], d['productos_asociados'], d['asunto'], d['observaciones'], d['fecha'], d['hora'], d['funcionario']))
        return res

    return render_template(
        'admin_documentos.html',
        actas=fmt_docs(actas_raw, 'actas'),
        informes=fmt_docs(informes_raw, 'informes'),
        reportes=fmt_docs(reportes_raw, 'reportes'),
        comisiones=fmt_docs(comisiones_raw, 'comisiones'),
        page=page,
        per_page=per_page,
        pages_a=pages_a, pages_i=pages_i, pages_r=pages_r, pages_c=pages_c,
        empresas=get_all_companies(),
        empresa_sel=empresa_filtro,
        active_tab=active_tab
    )

@app.route('/admin/editar/<tipo>/<int:id>', methods=['GET', 'POST'])
def editar_documento(tipo, id):
    if not require_login():
        return redirect('/')

    # Verificar si es admin o si el documento le pertenece
    is_admin = session.get('rol') == 'admin'
    
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute(f"SELECT id_usuario FROM {tipo} WHERE id=%s", (id,))
        row = cur.fetchone()
        if not row:
            flash("Documento no encontrado.", "warning")
            return redirect('/dashboard')
        
        if not is_admin and row[0] != session.get('user_id'):
            flash("No tienes permiso para editar este documento.", "danger")
            return redirect('/mis_documentos')

    TABLAS = {
        'actas': 'actas',
        'informes': 'informes',
        'reportes': 'reportes',
        'comisiones': 'comisiones'
    }

    # 1. Obtener catálogos (Listas)
    empresas = get_catalogo('EMPRESA')
    
    # Inicializar otras listas vacías por defecto
    gestiones = []
    productos = []
    report_types_json = []
    tipos_reporte = []

    if tipo == 'informes':
        gestiones = get_catalogo('GESTION_INFORME')
        productos = get_catalogo('PRODUCTO_INFORME')
        report_types_json = get_jerarquia_informes()
    elif tipo == 'reportes':
        gestiones = get_catalogo('GESTION_REPORTE')
        productos = get_catalogo('PRODUCTO_REPORTE')
        tipos_reporte = get_catalogo('TIPO_REPORTE')
    
    elif tipo == 'actas':
        gestiones = get_catalogo('GESTION_ACTA')
        productos = get_catalogo('PRODUCTO_ACTA')

    elif tipo == 'comisiones':
        gestiones = get_catalogo('GESTION_COMISION')
        productos = get_catalogo('PRODUCTO_COMISION')

    tipo = (tipo or "").lower()
    tabla = TABLAS.get(tipo)
    if not tabla:
        flash("Tipo de documento inválido.", "danger")
        return redirect('/admin/documentos')

    def empty_to_none(x):
        x = (x or "").strip()
        return x if x else None

    def resolver_otro(val, catalogo):
        if not val:
            return "", ""
        # Excluimos "Otros" de las opciones estándar para que cualquier valor 
        # (incluyendo el texto "Otros" si se guardó mal) se trate como valor personalizado.
        opciones = [(c.get('valor') or c.get('nombre')) for c in catalogo 
                    if (c.get('valor') or c.get('nombre')) != "Otros"]
        
        if val in opciones:
            return val, ""
        
        # Si no es una opción estándar, es un valor personalizado ("Otros" + el texto)
        return "Otros", val

    # POST -> actualizar
    if request.method == 'POST':
        numero_manual = request.form.get('numero_manual', '').strip()
        try:
            numero_manual = int(numero_manual) if numero_manual else None
        except:
            numero_manual = None
        
        asunto = request.form.get('asunto', '').strip()
        observaciones = request.form.get('observaciones', '').strip()
        empresa = request.form.get('empresa', '').strip()
        empresa_otro = request.form.get('empresa_otro', '').strip()

        if empresa == "Otros":
            if not empresa_otro:
                flash("Debe especificar la empresa cuando selecciona 'Otros'.", "warning")
                return redirect(request.url)
            empresa = empresa_otro

        # Informes/Reportes: soportar "Otros"
        tipo_informe = request.form.get('tipo_informe', '').strip()
        tipo_informe_otro = request.form.get('tipo_informe_otro', '').strip()

        tipo_reporte = request.form.get('tipo_reporte', '').strip()
        tipo_reporte_otro = request.form.get('tipo_reporte_otro', '').strip()

        # ===============================
        # INFORMES: gestiones / productos_asociados
        # ===============================
        gestiones_val = request.form.get('gestiones', '').strip()
        gestiones_otro = request.form.get('gestiones_otro', '').strip()

        productos_asociados = request.form.get('productos_asociados', '').strip()
        productos_asociados_otro = request.form.get('productos_asociados_otro', '').strip()

        # ===============================
        # REPORTES: gestiones / productos_asociados (NUEVO)
        # ===============================
        gestiones_reporte = request.form.get('gestiones_reporte', '').strip()
        gestiones_reporte_otro = request.form.get('gestiones_reporte_otro', '').strip()

        productos_asociados_reporte = request.form.get('productos_asociados_reporte', '').strip()
        productos_asociados_reporte_otro = request.form.get('productos_asociados_reporte_otro', '').strip()

        # ===============================
        # CAMPOS CASO FORTUITO (solo informes)
        # ===============================
        caso_tipo = request.form.get('caso_tipo', '').strip()
        nombre_alimentador = request.form.get('nombre_alimentador', '').strip()
        alimentador_subestacion = request.form.get('alimentador_subestacion', '').strip()
        linea_subtransmision_nombre = request.form.get('linea_subtransmision_nombre', '').strip()

        fecha_interrupcion = request.form.get('fecha_interrupcion', '').strip()

        if not asunto:
            flash("El asunto es obligatorio.", "warning")
            return redirect(request.url)

        if not empresa:
            flash("Debe seleccionar una empresa.", "warning")
            return redirect(request.url)

        # ===============================
        # Resolver "Otros" y validar por tipo
        # ===============================
        if tipo == 'informes':
            if not tipo_informe:
                flash("Debe seleccionar el tipo de informe.", "warning")
                return redirect(request.url)

            if tipo_informe == "Otros":
                if not tipo_informe_otro:
                    flash("Debe especificar el tipo de informe cuando selecciona 'Otros'.", "warning")
                    return redirect(request.url)
                tipo_informe = tipo_informe_otro

            if not gestiones_val:
                flash("Debe seleccionar una gestión.", "warning")
                return redirect(request.url)
            if gestiones_val == "Otros":
                if not gestiones_otro:
                    flash("Debe especificar la gestión cuando selecciona 'Otros'.", "warning")
                    return redirect(request.url)
                gestiones_val = gestiones_otro

            if not productos_asociados:
                flash("Debe seleccionar un producto asociado.", "warning")
                return redirect(request.url)
            if productos_asociados == "Otros":
                if not productos_asociados_otro:
                    flash("Debe especificar el producto asociado cuando selecciona 'Otros'.", "warning")
                    return redirect(request.url)
                productos_asociados = productos_asociados_otro

        elif tipo == 'reportes':
            if not tipo_reporte:
                flash("Debe seleccionar el tipo de reporte.", "warning")
                return redirect(request.url)

            if tipo_reporte == "Otros":
                if not tipo_reporte_otro:
                    flash("Debe especificar el tipo de reporte cuando selecciona 'Otros'.", "warning")
                    return redirect(request.url)
                tipo_reporte = tipo_reporte_otro

            # NUEVO: gestiones / productos_asociados obligatorios en reportes
            if not gestiones_reporte:
                flash("Debe seleccionar una gestión (Reportes).", "warning")
                return redirect(request.url)
            if gestiones_reporte == "Otros":
                if not gestiones_reporte_otro:
                    flash("Debe especificar la gestión cuando selecciona 'Otros' (Reportes).", "warning")
                    return redirect(request.url)
                gestiones_reporte = gestiones_reporte_otro

            if not productos_asociados_reporte:
                flash("Debe seleccionar un producto asociado (Reportes).", "warning")
                return redirect(request.url)
            if productos_asociados_reporte == "Otros":
                if not productos_asociados_reporte_otro:
                    flash("Debe especificar el producto asociado cuando selecciona 'Otros' (Reportes).", "warning")
                    return redirect(request.url)
                productos_asociados_reporte = productos_asociados_reporte_otro

        elif tipo in ['actas', 'comisiones']:
            if not gestiones_val:
                flash("Debe seleccionar una gestión.", "warning")
                return redirect(request.url)
            if gestiones_val == "Otros":
                if not gestiones_otro:
                    flash("Debe especificar la gestión cuando selecciona 'Otros'.", "warning")
                    return redirect(request.url)
                gestiones_val = gestiones_otro
            
            if not productos_asociados:
                flash("Debe seleccionar un producto asociado.", "warning")
                return redirect(request.url)
            if productos_asociados == "Otros":
                if not productos_asociados_otro:
                    flash("Debe especificar el producto asociado cuando selecciona 'Otros'.", "warning")
                    return redirect(request.url)
                productos_asociados = productos_asociados_otro
        # ===============================
        # Validación / limpieza Caso Fortuito (solo informes)
        # ===============================
        if tipo == 'informes':
            if "caso fortuito" in (tipo_informe or "").lower():
                if not caso_tipo:
                    flash("En 'Caso Fortuito' debe seleccionar: ALIMENTADOR o LINEAS DE SUBTRANSMISION.", "warning")
                    return redirect(request.url)

                if not fecha_interrupcion:
                    flash("En 'Caso Fortuito' debe seleccionar la FECHA de interrupción.", "warning")
                    return redirect(request.url)

                if caso_tipo == "ALIMENTADOR":
                    if not nombre_alimentador:
                        flash("En 'ALIMENTADOR' debe ingresar el nombre del alimentador.", "warning")
                        return redirect(request.url)
                    if not alimentador_subestacion:
                        flash("En 'ALIMENTADOR' debe ingresar la subestación.", "warning")
                        return redirect(request.url)

                    # limpiar campos del otro tipo
                    linea_subtransmision_nombre = ""

                elif caso_tipo == "LINEAS DE SUBTRANSMISION":
                    if not linea_subtransmision_nombre:
                        flash("En 'LINEAS DE SUBTRANSMISION' debe ingresar el nombre.", "warning")
                        return redirect(request.url)

                    # limpiar campos del otro tipo
                    nombre_alimentador = ""
                    alimentador_subestacion = ""

                else:
                    flash("Selección inválida en Caso Tipo.", "warning")
                    return redirect(request.url)

            else:
                caso_tipo = ""
                nombre_alimentador = ""
                alimentador_subestacion = ""
                linea_subtransmision_nombre = ""
                fecha_interrupcion = ""

        # ===============================
        # UPDATE
        # ===============================
        with get_conn() as conn, conn.cursor() as cur:
            if tipo == 'informes':
                cur.execute("""
                    UPDATE informes
                    SET
                        numero=COALESCE(%s, numero),
                        empresa=%s,
                        tipo_informe=%s,

                        gestiones=%s,
                        productos_asociados=%s,

                        asunto=%s,
                        observaciones=%s,

                        caso_tipo=%s,
                        nombre_alimentador=%s,
                        alimentador_subestacion=%s,
                        linea_subtransmision_nombre=%s,
                        fecha_interrupcion=%s
                    WHERE id=%s
                """, (
                    empresa,
                    tipo_informe,

                    empty_to_none(gestiones_val),
                    empty_to_none(productos_asociados),

                    asunto,
                    observaciones,

                    empty_to_none(caso_tipo),
                    empty_to_none(nombre_alimentador),
                    empty_to_none(alimentador_subestacion),
                    empty_to_none(linea_subtransmision_nombre),
                    empty_to_none(fecha_interrupcion),
                    id
                ))

            elif tipo == 'reportes':
                cur.execute("""
                    UPDATE reportes
                    SET
                        numero=COALESCE(%s, numero),
                        empresa=%s,
                        gestiones=%s,
                        productos_asociados=%s,
                        tipo_reporte=%s,
                        asunto=%s,
                        observaciones=%s
                    WHERE id=%s
                """, (
                    numero_manual,
                    empresa,
                    empty_to_none(gestiones_reporte),
                    empty_to_none(productos_asociados_reporte),
                    tipo_reporte,
                    asunto,
                    observaciones,
                    id
                ))

            else:
                cur.execute(f"""
                    UPDATE {tabla}
                    SET
                        numero=COALESCE(%s, numero),
                        empresa=%s,
                        asunto=%s,
                        observaciones=%s,
                        gestiones=%s,
                        productos_asociados=%s
                    WHERE id=%s
                """, (
                    numero_manual,
                    empresa,
                    asunto,
                    observaciones,
                    empty_to_none(gestiones_val),
                    empty_to_none(productos_asociados),
                    id
                ))

            conn.commit()

        flash(f"{tipo.capitalize()} actualizado correctamente.", "success")
        return redirect('/admin/documentos')

    with get_conn() as conn, conn.cursor() as cur:
        if tipo == 'informes':
            cur.execute("""
                SELECT
                    numero,
                    empresa,
                    tipo_informe,
                    gestiones,
                    productos_asociados,
                    asunto,
                    observaciones,
                    caso_tipo,
                    nombre_alimentador,
                    alimentador_subestacion,
                    linea_subtransmision_nombre,
                    fecha_interrupcion
                FROM informes
                WHERE id=%s
            """, (id,))
            doc = cur.fetchone()

            if not doc:
                flash("Documento no encontrado.", "warning")
                return redirect('/admin/documentos')

            return render_template(
                'formulario.html',
                tipo=tipo,
                editar=True,
                id=id,
                
                # Catalogs (Lists)
                empresas=empresas,
                gestiones=gestiones,
                productos_asociados=productos,
                report_types_json=report_types_json,
                
                # Valores seleccionados (convención *_sel para evitar colisiones)
                numero_original=doc[0],
                empresa_sel=resolver_otro(doc[1], empresas)[0],
                empresa_otro=resolver_otro(doc[1], empresas)[1],
                
                # Tipo Informe con manejo de Otros
                tipo_informe_sel=resolver_otro(doc[2], get_catalogo('TIPO_INFORME'))[0],
                tipo_informe_otro=resolver_otro(doc[2], get_catalogo('TIPO_INFORME'))[1],
                
                # Manejo de Otros
                gestiones_sel=resolver_otro(doc[3], gestiones)[0],
                gestiones_otro=resolver_otro(doc[3], gestiones)[1],
                
                productos_asociados_sel=resolver_otro(doc[4], productos)[0],
                productos_asociados_otro=resolver_otro(doc[4], productos)[1],
                
                asunto=doc[5],
                observaciones=doc[6],

                # Special Fields
                caso_tipo=doc[7] or "",
                nombre_alimentador=doc[8] or "",
                alimentador_subestacion=doc[9] or "",
                linea_subtransmision_nombre=doc[10] or "",
                
                fecha_interrupcion=doc[11].strftime("%Y-%m-%d") if doc[11] else ""
            )

        elif tipo == 'reportes':
            cur.execute("""
                SELECT
                    numero,
                    empresa,
                    gestiones,
                    productos_asociados,
                    tipo_reporte,
                    asunto,
                    observaciones
                FROM reportes
                WHERE id=%s
            """, (id,))
            doc = cur.fetchone()

            if not doc:
                flash("Documento no encontrado.", "warning")
                return redirect('/admin/documentos')

            return render_template(
                'formulario.html',
                tipo=tipo,
                editar=True,
                id=id,
                
                # Catálogos
                empresas=empresas,
                gestiones=gestiones,
                productos_asociados=productos,
                tipos_reporte=tipos_reporte,
                report_types_json=[],
                causas_caso_fortuito=[],


                # Valores seleccionados con manejo de Otros
                productos_asociados_sel=resolver_otro(doc[2], productos)[0],
                productos_asociados_otro=resolver_otro(doc[2], productos)[1],
                
                tipo_reporte_sel=resolver_otro(doc[3], tipos_reporte)[0],
                tipo_reporte_otro=resolver_otro(doc[3], tipos_reporte)[1],
                
                asunto=doc[4],
                observaciones=doc[5]
            )

        else:
            # ACTAS / COMISIONES
            cur.execute(f"""
                SELECT numero, empresa, asunto, observaciones, gestiones, productos_asociados
                FROM {tabla}
                WHERE id=%s
            """, (id,))
            doc = cur.fetchone()

            if not doc:
                flash("Documento no encontrado.", "warning")
                return redirect('/admin/documentos')

            return render_template(
                'formulario.html',
                tipo=tipo,
                editar=True,
                id=id,
                
                # Catalogs
                empresas=empresas,
                gestiones=gestiones,
                productos_asociados=productos,
                causas_caso_fortuito=[], # Vacío para actas/comisiones
                
                # Valores seleccionados

                numero_original=doc[0],
                empresa_sel=resolver_otro(doc[1], empresas)[0],
                empresa_otro=resolver_otro(doc[1], empresas)[1],
                asunto=doc[2],
                observaciones=doc[3],
                
                # Manejo de Otros
                gestiones_sel=resolver_otro(doc[4], gestiones)[0],
                gestiones_otro=resolver_otro(doc[4], gestiones)[1],
                
                productos_asociados_sel=resolver_otro(doc[5], productos)[0],
                productos_asociados_otro=resolver_otro(doc[5], productos)[1]
            )



@app.route('/eliminar/<tipo>/<int:id>')
def eliminar_documento(tipo, id):
    if not require_admin():
        return redirect('/')

    tipo = (tipo or "").lower()
    if tipo not in ['actas', 'informes', 'reportes', 'comisiones']:
        flash("Tipo de documento inválido.", "danger")
        return redirect('/admin/documentos')

    with get_conn() as conn, conn.cursor() as cur:
        cur.execute(f"DELETE FROM {tipo} WHERE id = %s", (id,))
        conn.commit()

    flash("Documento eliminado correctamente.", "danger")
    return redirect('/admin/documentos')

# ===============================
# CRUD usuarios (admin)
# ===============================
@app.route('/admin/usuarios/crear', methods=['GET', 'POST'])
def crear_usuario():
    if not require_admin():
        return redirect('/')

    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        rol = request.form.get('rol', 'usuario')

        with get_conn() as conn, conn.cursor() as cur:
            cur.execute("""
                INSERT INTO usuarios (nombre, email, password, rol)
                VALUES (%s, %s, %s, %s)
            """, (nombre, email, password, rol))  # Se guarda en texto plano
            conn.commit()


        flash("Usuario creado exitosamente", "success")
        return redirect('/admin')

    return render_template('form_usuario.html', modo="Crear", usuario=None)

# ===============================
# Editar usuario (admin)
# ===============================
@app.route('/admin/usuarios/editar/<int:id>', methods=['GET', 'POST'])
def editar_usuario(id):
    if not require_admin():
        return redirect('/')

    with get_conn() as conn, conn.cursor() as cur:

        # POST -> actualizar
        if request.method == 'POST':
            nombre = request.form.get('nombre', '').strip()
            email = request.form.get('email', '').strip().lower()
            rol = request.form.get('rol', 'usuario').strip()

            # opcional: cambiar password solo si se escribió algo
            password = request.form.get('password', '').strip()

            if not nombre or not email:
                flash("Nombre y correo son obligatorios.", "warning")
                return redirect(request.url)

            if rol not in ['usuario', 'admin']:
                flash("Rol inválido.", "warning")
                return redirect(request.url)

            # Validar email único (que no sea de otro usuario)
            cur.execute("""
                SELECT 1 FROM usuarios
                WHERE email=%s AND id<>%s
            """, (email, id))
            if cur.fetchone():
                flash("Ese correo ya está registrado en otro usuario.", "warning")
                return redirect(request.url)

            if password:
                cur.execute("""
                    UPDATE usuarios
                    SET nombre=%s, email=%s, rol=%s, password=%s
                    WHERE id=%s
                """, (nombre, email, rol, password, id))
            else:
                cur.execute("""
                    UPDATE usuarios
                    SET nombre=%s, email=%s, rol=%s
                    WHERE id=%s
                """, (nombre, email, rol, id))

            conn.commit()
            flash("Usuario actualizado correctamente.", "success")
            return redirect('/admin')

        # GET -> cargar usuario
        cur.execute("""
            SELECT id, nombre, email, rol
            FROM usuarios
            WHERE id=%s
        """, (id,))
        usuario = cur.fetchone()

    if not usuario:
        flash("Usuario no encontrado.", "warning")
        return redirect('/admin')

    return render_template('form_usuario.html', modo="Editar", usuario=usuario)

@app.route('/admin/usuarios/eliminar/<int:id>')
def eliminar_usuario(id):
    if not require_admin():
        return redirect('/')

    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("DELETE FROM usuarios WHERE id=%s", (id,))
        conn.commit()

    flash("Usuario eliminado", "danger")
    return redirect('/admin')

# ===============================
# Listados para usuario
# ===============================
@app.route('/mis_documentos')
def mis_documentos():
    if not require_login():
        return redirect('/')

    per_page = request.args.get('per_page', '10')
    page = request.args.get('page', 1, type=int)
    if page < 1: page = 1
    empresa_filtro = request.args.get('empresa', '').strip()
    active_tab = request.args.get('tab', 'actas')
    
    # Comportamiento original: ver TODO (o corregir si se desea filtrar por user)
    # user_id = session['user_id'] # Descomentar para filtrar por usuario
    user_id = None

    actas_raw, total_a, pages_a = get_paginated_docs('actas', empresa_filtro, user_id=user_id, page=page, per_page=per_page)
    informes_raw, total_i, pages_i = get_paginated_docs('informes', empresa_filtro, user_id=user_id, page=page, per_page=per_page)
    reportes_raw, total_r, pages_r = get_paginated_docs('reportes', empresa_filtro, user_id=user_id, page=page, per_page=per_page)
    comisiones_raw, total_c, pages_c = get_paginated_docs('comisiones', empresa_filtro, user_id=user_id, page=page, per_page=per_page)

    def fmt_docs(docs, tipo):
        res = []
        for d in docs:
            codigo = format_doc_code(tipo, d['numero'], d['anio'])
            if tipo == 'informes':
                res.append((d['id'], codigo, d['empresa'], d['gestiones'], d['productos_asociados'], d['tipo_informe'], d['caso_tipo'], d['nombre_alimentador'], d['alimentador_subestacion'], d['linea_subtransmision_nombre'], d['fecha_interrupcion'], d['asunto'], d['observaciones'], d['fecha'], d['hora'], d['funcionario']))
            elif tipo == 'reportes':
                res.append((d['id'], codigo, d['empresa'], d['gestiones'], d['productos_asociados'], d['tipo_reporte'], d['asunto'], d['observaciones'], d['fecha'], d['hora'], d['funcionario']))
            else:
                res.append((d['id'], codigo, d['empresa'], d['gestiones'], d['productos_asociados'], d['asunto'], d['observaciones'], d['fecha'], d['hora'], d['funcionario']))
        return res

    return render_template(
        'mis_documentos.html',
        actas=fmt_docs(actas_raw, 'actas'),
        informes=fmt_docs(informes_raw, 'informes'),
        reportes=fmt_docs(reportes_raw, 'reportes'),
        comisiones=fmt_docs(comisiones_raw, 'comisiones'),
        page=page,
        per_page=per_page,
        pages_a=pages_a, pages_i=pages_i, pages_r=pages_r, pages_c=pages_c,
        empresas=get_all_companies(),
        empresa_sel=empresa_filtro,
        active_tab=active_tab
    )


# ===============================
# Utilidad CLI (opcional)
# ===============================
def crear_usuario_directo(nombre, email, password, rol):
    """Utilidad para crear un usuario desde consola usando la misma config."""
    with get_conn() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(
            "INSERT INTO usuarios (nombre, email, password, rol) VALUES (%s, %s, %s, %s) RETURNING *",
            (nombre, email.lower(), password, rol)
        )
        row = cur.fetchone()
        conn.commit()
        return dict(row)

# ===============================
# Principal
# ===============================
if __name__ == '__main__':
    # En desarrollo local: flask built-in; en Docker se usa Gunicorn (CMD del contenedor)
    app.run(host='0.0.0.0', port=8000, debug=True)
